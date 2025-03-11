import ftplib
import os
import argparse
import traceback
from scrapers_config import SCRAPERS, DAILY_SCRAPERS, MONTHLY_SCRAPERS

# Parse command line arguments
parser = argparse.ArgumentParser(description='Download files from FTP')
parser.add_argument('scraper', nargs='?', help='Specific scraper to process')
parser.add_argument('--type', choices=['daily', 'monthly'], default='daily',
                    help='Type of scrapers to process (daily or monthly)')
args = parser.parse_args()

# Ensure Pricing Spreadsheets directory exists
if not os.path.exists('Pricing Spreadsheets'):
    os.makedirs('Pricing Spreadsheets')

# Determine which scrapers to process
if args.scraper:
    if args.scraper not in SCRAPERS:
        print(f"Error: Scraper '{args.scraper}' not found in config")
        exit(1)
    scrapers_to_process = [args.scraper]
else:
    # Process all scrapers of the specified type
    if args.type == 'monthly':
        scrapers_to_process = MONTHLY_SCRAPERS.keys()
    else:  # default to daily
        scrapers_to_process = DAILY_SCRAPERS.keys()

print("Connecting to FTP for downloading files...")

# Get password from environment variable
password = os.environ.get('FTP_PASSWORD')
if not password:
    raise ValueError("FTP_PASSWORD environment variable not set")

try:
    # Connect to FTP
    session = ftplib.FTP('ftp.drivehq.com', 'kyaldabomb', password)
    
    # Check if directory exists
    if 'competitor_pricing' not in session.nlst():
        print("competitor_pricing directory not found on FTP server")
        exit(1)
    
    # Change to the directory
    session.cwd('competitor_pricing')
    
    # Get the list of files on the FTP server
    ftp_files = session.nlst()
    
    # Process each scraper
    for scraper_name in scrapers_to_process:
        scraper = SCRAPERS[scraper_name]
        file_name = scraper["file_name"]
        file_path = f'Pricing Spreadsheets/{file_name}'
        
        print(f"Looking for {file_name} on FTP server...")
        if file_name in ftp_files:
            print(f"Downloading {file_name}...")
            try:
                with open(file_path, 'wb') as file:
                    session.retrbinary(f'RETR {file_name}', file.write)
                print(f"Download of {file_name} complete")
            except Exception as download_error:
                print(f"Error downloading {file_name}: {str(download_error)}")
                print(traceback.format_exc())
                
                # If file doesn't exist locally, create a new Excel file
                if not os.path.exists(file_path):
                    try:
                        import openpyxl
                        print(f"Creating new Excel file for {file_name}")
                        wb = openpyxl.Workbook()
                        sheet = wb.active
                        sheet.title = "Sheet"
                        
                        # Add headers
                        headers = ["SKU", "Brand", "Title", "Price", "URL", "Image", "Description", "Last Updated", "In Stock"]
                        for i, header in enumerate(headers, 1):
                            sheet.cell(row=1, column=i).value = header
                        
                        wb.save(file_path)
                        print(f"Created new Excel file: {file_path}")
                    except Exception as create_error:
                        print(f"Error creating new Excel file: {str(create_error)}")
                        print(traceback.format_exc())
        else:
            print(f"File {file_name} not found on FTP server")
            
            # Create a new Excel file if it doesn't exist locally
            if not os.path.exists(file_path):
                try:
                    import openpyxl
                    print(f"Creating new Excel file for {file_name}")
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    sheet.title = "Sheet"
                    
                    # Add headers
                    headers = ["SKU", "Brand", "Title", "Price", "URL", "Image", "Description", "Last Updated", "In Stock"]
                    for i, header in enumerate(headers, 1):
                        sheet.cell(row=1, column=i).value = header
                    
                    wb.save(file_path)
                    print(f"Created new Excel file: {file_path}")
                except Exception as create_error:
                    print(f"Error creating new Excel file: {str(create_error)}")
                    print(traceback.format_exc())
    
    # Close the FTP connection
    session.quit()
    print("File downloads completed")
except Exception as e:
    print(f"Error during FTP download: {str(e)}")
    print(traceback.format_exc())
    exit(1)
