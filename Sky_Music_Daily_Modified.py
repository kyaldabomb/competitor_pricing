import openpyxl
import requests
from bs4 import BeautifulSoup
import re, math
from requests_html import HTMLSession
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium_stealth import stealth
import os, time, traceback
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import sys
import argparse
import ftplib


# Parse command line arguments
parser = argparse.ArgumentParser(description='Run daily scraper for Sky Music')
parser.add_argument('scraper', nargs='?', default='sky_music_daily', 
                    help='Scraper name from config')
args = parser.parse_args()
def upload_to_ftp(file_path, file_name):
    print(f"\n==== Uploading {file_name} to FTP ====")
    try:
        # Get FTP password from environment
        ftp_password = os.environ.get('FTP_PASSWORD')
        if not ftp_password:
            print("FTP_PASSWORD not found in environment variables")
            return False
            
        # Connect and upload
        session = ftplib.FTP('ftp.drivehq.com', 'kyaldabomb', ftp_password)
        
        # Check if directory exists
        if 'competitor_pricing' not in session.nlst():
            print("competitor_pricing directory not found, creating it...")
            session.mkd('competitor_pricing')
        
        # Change to the directory
        session.cwd('competitor_pricing')
        
        # Upload the file
        with open(file_path, 'rb') as file:
            session.storbinary(f'STOR {file_name}', file)
            
        # Create timestamp file for verification
        with open('upload_timestamp.txt', 'w') as f:
            f.write(f"Upload of {file_name} completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
        with open('upload_timestamp.txt', 'rb') as file:
            session.storbinary('STOR upload_timestamp.txt', file)
            
        session.quit()
        print(f"File {file_name} uploaded to FTP successfully")
        return True
    except Exception as e:
        print(f"Error uploading to FTP: {str(e)}")
        print(traceback.format_exc())
        return False
      
# Email notification function
def send_email_notification(success, items_count=0, error_msg=""):
    print("Sending email notification...")
    try:
        # Email settings
        sender = "kyal@scarlettmusic.com.au"
        receiver = "kyal@scarlettmusic.com.au"
        password = os.environ.get('EMAIL_PASSWORD')
        if not password:
            print("Email password not found in environment variables")
            return
            
        host = "mail.scarlettmusic.com.au"
        port = 587
        
        msg = MIMEMultipart()
        msg['From'] = sender
        msg['To'] = receiver
        
        if success:
            msg['Subject'] = f"Sky Music Daily Scraper Success: {items_count} items updated"
            body = f"The Sky Music daily web scraper ran successfully and updated {items_count} items."
        else:
            msg['Subject'] = "Sky Music Daily Scraper Failed"
            body = f"The Sky Music daily web scraper encountered an error:\n\n{error_msg}"
        
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(host, port)
        server.starttls()
        server.login(sender, password)
        server.send_message(msg)
        server.quit()
        print("Email notification sent successfully")
    except Exception as e:
        print(f"Failed to send email notification: {str(e)}")
        print(traceback.format_exc())

# Setup Chrome options for headless operation in GitHub Actions
options = webdriver.ChromeOptions()
options.add_argument("--headless")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)

try:
    # Initialize WebDriver using webdriver-manager
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    # Print Chrome and ChromeDriver version for debugging
    print(f"Chrome version: {driver.capabilities['browserVersion']}")
    print(f"ChromeDriver version: {driver.capabilities['chrome']['chromedriverVersion'].split(' ')[0]}")
    
    stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
            )
    
    session = HTMLSession()
    
    # Use local path instead of network path
    file_path = "Pricing Spreadsheets/Sky_Music.xlsx"
    file_name = "Sky_Music.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet']
    
    item_number = 0
    items_scrapped = 0
    
    print(f"Starting to process {sheet.max_row-1} items")
    
    for sheet_line in range(2, sheet.max_row+1):
        try:
            item_number += 1

            time_last_scrapped = sheet['H' + str(sheet_line)].value
            if time_last_scrapped:
                try:
                    string_datetime_conversion = datetime.strptime(time_last_scrapped, '%m %d %Y')
                    if string_datetime_conversion + timedelta(days=3) > datetime.today():
                        print(f'Item {str(item_number)} scrapped less than 7 days ago, skipping...')
                        continue
                except Exception as date_error:
                    print(f"Error parsing date {time_last_scrapped}: {str(date_error)}")
            
            # Get the URL and fetch the page
            url = sheet['E'+str(sheet_line)].value
            if not url:
                print(f"No URL for item {item_number}, skipping")
                continue
                
            print(f"Processing item {item_number}: {url}")
            
            # Add retry logic for resilience
            max_retries = 3
            for retry in range(max_retries):
                try:
                    r = driver.get(url)
                    break
                except Exception as e:
                    if retry == max_retries - 1:
                        raise
                    print(f"Retry {retry+1}/{max_retries} for {url}: {str(e)}")
                    time.sleep(5)
            
            html = driver.page_source
            soup2 = BeautifulSoup(html, features="lxml")
            
            # Extract brand
            try:
                brand = soup2.find(class_='vendor').text.strip()
            except:
                print(f"Could not find brand for {url}, skipping")
                continue
                
            print(f"Brand: {brand}")
            
            # Extract SKU
            try:
                sku = soup2.find(class_='sku').text.strip()
            except:
                print(f"Could not find SKU for {url}, skipping")
                continue
            
            # Special handling for certain brands
            if brand == 'Ernie Ball':
                sku = sku.replace('P0', '')
            
            if brand.lower() == 'orange':
                sku = f'{sku}AUSTRALIS'
            
            # Extract title
            title = soup2.find(class_='product_name').text.strip()
            
            # Extract price
            try:
                price = soup2.find(class_='price price--sale').text.strip()
                price = price.replace('$', '')
                price = price.replace(',', '')
            except:
                try:
                    price = soup2.find(class_='compare-at-price').text.strip()
                    price = price.replace('$', '')
                    price = price.replace(',', '')
                except:
                    price = "N/A"
                    print(f"Could not find price for {url}")
            
            print(f'\nScraping Item {str(item_number)}\nSKU: {sku}\nPrice: {price}\n')
            
            # Extract image
            image = 'Not yet scraped'
            try:
                image_element = soup2.find(class_='gallery-cell is-selected')
                if image_element:
                    image_link = image_element.find('a')
                    if image_link and 'href' in image_link.attrs:
                        image = image_link['href']
            except Exception as e:
                print(f"Error getting image: {str(e)}")
            
            # Extract description
            description = 'Not yet scraped'
            try:
                description = soup2.find(class_='station-tabs-content-inner').text
            except:
                description = 'N/A'
            
            # Check stock availability
            stock_avaliable = 'n'
            for x in soup2.find_all(class_='location-stock-status'):
                if 'In Stock' in x.text or "Low Stock" in x.text:
                    stock_avaliable = 'y'
                    break
            try:
              if 'In Stock' in soup.find(class_='iia-location-info').text:
                stock_avaliable = 'y'
            except:
                pass
            
            # Get current date
            today = datetime.now()
            date = today.strftime('%m %d %Y')
            
            # Update the Excel sheet
            sheet['A' + str(sheet_line)].value = sku
            sheet['B' + str(sheet_line)].value = brand
            sheet['C' + str(sheet_line)].value = title
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliable
            
            items_scrapped += 1
            print(f'Item {str(item_number)} scraped successfully')
            
            # Add a pause to be gentle with the server
            time.sleep(5)
            
            # Save periodically
            if int(items_scrapped) % 100 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(file_path)
                    upload_to_ftp(file_path, file_name)

                    print("Sheet saved successfully")
                except Exception as e:
                    print(f"Error occurred while saving the Excel file: {str(e)}")
        
        except Exception as item_error:
            print(f"Error processing item {item_number}: {str(item_error)}")
            print(traceback.format_exc())
            # Continue with next item even if this one fails
    
    # Final save
    wb.save(file_path)
    upload_to_ftp(file_path, file_name)

    print(f"Scraping completed successfully. Updated {items_scrapped} items.")
    send_email_notification(True, items_scrapped)
    
except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")
    
    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")
    
    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)  # Exit with error code
finally:
    # Always close the driver
    try:
        driver.quit()
    except:
        pass
