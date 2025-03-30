import openpyxl
import requests
from bs4 import BeautifulSoup
import os
import time
import traceback
from datetime import datetime, timedelta
import sys
import argparse

# Import helper functions (assuming they are in the same directory or accessible)
try:
    from ftp_helper import upload_to_ftp
    from email_notifications import send_email_notification
except ImportError:
    print("Error: ftp_helper.py or email_notifications.py not found.")
    print("Ensure these files are in the same directory or Python path.")
    sys.exit(1)

# --- Configuration and Argument Parsing ---

# Parse command line arguments
parser = argparse.ArgumentParser(description='Run daily scraper for Australian Piano Warehouse (APW)')
parser.add_argument('scraper', nargs='?', default='apw_daily',
                    help='Scraper name (should be apw_daily)')
args = parser.parse_args()
scraper_name = args.scraper

# Basic check if the scraper name matches
if scraper_name != "apw_daily":
    print(f"Warning: Expected scraper name 'apw_daily', but received '{scraper_name}'. Proceeding anyway.")

# Define file names and paths (using relative paths for Actions)
file_name = "APW.xlsx"
file_path = f"Pricing Spreadsheets/{file_name}"
description = "Australian Piano Warehouse (Daily)" # Description for notifications

# --- Main Script Logic ---

wb = None # Initialize wb to None for robust error handling in finally block
try:
    print(f"Starting {description} scraper...")
    print(f"Loading workbook: {file_path}")

    # Check if the file exists before trying to load
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}. Please ensure the file was downloaded correctly.")
        # Optionally create a blank file here if needed for the first run,
        # but the download_files.py script should handle this.
        sys.exit(1)

    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet']

    item_number = 0
    items_scrapped = 0
    total_items = sheet.max_row - 1

    print(f"Starting to process {total_items} items from {file_name}")

    for sheet_line in range(2, sheet.max_row + 1):
        item_number += 1
        current_item_info = f"Item {item_number}/{total_items} (Sheet Row {sheet_line})"

        try:
            # --- Date Check ---
            time_last_scrapped_val = sheet['H' + str(sheet_line)].value
            skip_item = False
            if isinstance(time_last_scrapped_val, datetime):
                # If it's already a datetime object (less likely from direct Excel read unless formatted perfectly)
                time_last_scrapped = time_last_scrapped_val
                if time_last_scrapped + timedelta(days=7) > datetime.now():
                    print(f'{current_item_info}: Scraped within last 7 days ({time_last_scrapped.strftime("%Y-%m-%d")}), skipping...')
                    skip_item = True
            elif isinstance(time_last_scrapped_val, str) and time_last_scrapped_val.strip():
                # If it's a string, try to parse it
                try:
                    # Adjust the format string '%m %d %Y' to match exactly how dates are stored
                    time_last_scrapped = datetime.strptime(time_last_scrapped_val.strip(), '%m %d %Y')
                    if time_last_scrapped + timedelta(days=7) > datetime.now():
                        print(f'{current_item_info}: Scraped within last 7 days ({time_last_scrapped.strftime("%Y-%m-%d")}), skipping...')
                        skip_item = True
                except ValueError as date_error:
                    print(f"{current_item_info}: Error parsing date '{time_last_scrapped_val}'. Proceeding with scrape. Error: {date_error}")
            else:
                # If it's empty or None, proceed with scrape
                print(f"{current_item_info}: No valid last scraped date found, proceeding with scrape.")

            if skip_item:
                continue

            # --- Get URL and Fetch Data ---
            url = sheet['E' + str(sheet_line)].value
            if not url or not url.startswith(('http://', 'https://')):
                print(f"{current_item_info}: Invalid or missing URL ('{url}'), skipping.")
                continue

            print(f"{current_item_info}: Processing URL: {url}")

            # Add retry logic for requests
            max_retries = 3
            response = None
            for attempt in range(max_retries):
                try:
                    response = requests.get(url, timeout=30) # Add timeout
                    response.raise_for_status() # Raise an exception for bad status codes (4xx or 5xx)
                    break # Success
                except requests.exceptions.RequestException as req_err:
                    print(f"{current_item_info}: Request attempt {attempt + 1}/{max_retries} failed: {req_err}")
                    if attempt == max_retries - 1:
                        print(f"{current_item_info}: Max retries reached. Skipping item.")
                        raise # Re-raise the last exception to be caught by the outer try/except
                    time.sleep(5 * (attempt + 1)) # Exponential backoff

            if response is None: # Should not happen if raise works, but as safeguard
                continue

            soup = BeautifulSoup(response.content, 'html.parser') # Use html.parser or lxml if installed

            # --- Extract Data ---
            title = "N/A"
            price = "N/A"
            image = "N/A"
            description_text = "N/A"
            sku = "N/A"
            stock_availability = 'n' # Default to 'n'

            # Title
            title_tag = soup.find(class_='product_title entry-title wd-entities-title')
            if title_tag:
                title = title_tag.text.strip()
            else:
                print(f"{current_item_info}: Could not find title tag.")
                # Consider skipping if title is essential: continue

            # Price
            price_tag = soup.find('p', class_='price')
            if price_tag:
                # Handle potential variations in price display (e.g., sale price, regular price)
                # This example takes the first price found. More complex logic might be needed.
                price_text = price_tag.text.strip()
                # Clean price: remove $, ,, whitespace. Handle potential ranges or multiple prices if needed.
                cleaned_price = price_text.replace('$', '').replace(',', '').split()[0] # Take first part if space separated
                try:
                    # Verify it's a number-like value before assigning
                    float(cleaned_price)
                    price = cleaned_price
                except ValueError:
                    print(f"{current_item_info}: Found price text ('{price_text}') but failed to clean to a number.")
                    price = "N/A" # Keep as N/A if cleaning fails
            else:
                print(f"{current_item_info}: Could not find price tag.")

            # Image
            image_tag = soup.find(class_='attachment-woocommerce_single size-woocommerce_single wp-post-image')
            if image_tag and 'src' in image_tag.attrs:
                image = image_tag['src']
            else:
                 print(f"{current_item_info}: Could not find image tag.")

            # Description
            desc_tag = soup.find(class_='wc-tab-inner') # Or check 'col-12 poduct-tabs-inner' if needed
            if desc_tag:
                description_text = desc_tag.text.strip()
            else:
                 print(f"{current_item_info}: Could not find description tag.")

            # SKU
            sku_tag = soup.find(class_='sku')
            if sku_tag:
                sku = sku_tag.text.strip()
            else:
                 print(f"{current_item_info}: Could not find SKU tag.")


            # Stock
            stock_tag = soup.find(class_='stock-feeds-stock') # Check if this class is reliable
            if stock_tag:
                stock_text = stock_tag.text.lower()
                if 'in stock' in stock_text:
                    stock_availability = 'y'
                # else: stock_availability remains 'n' (default)
            else:
                # Fallback: Check for general 'out of stock' messages if specific tag is missing
                out_of_stock_tag = soup.find(class_='out-of-stock') # Example class, adjust as needed
                if not out_of_stock_tag:
                     # If no specific stock tag and no 'out-of-stock' tag found, assume 'in stock'?
                     # This assumption might be wrong. Defaulting to 'n' is safer unless confirmed otherwise.
                     print(f"{current_item_info}: Could not find specific stock tag. Defaulting stock to 'n'.")
                     # stock_availability = 'y' # Uncomment cautiously if this is the desired logic

            # --- Update Spreadsheet ---
            today = datetime.now()
            date_str = today.strftime('%m %d %Y') # Format consistent with existing data

            sheet['C' + str(sheet_line)].value = title
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description_text
            sheet['H' + str(sheet_line)].value = date_str
            sheet['I' + str(sheet_line)].value = stock_availability
            # Ensure 'A' (SKU) and 'B' (Brand - if applicable) are also updated if needed
            sheet['A' + str(sheet_line)].value = sku
            # sheet['B' + str(sheet_line)].value = brand # Add brand extraction if possible/needed

            items_scrapped += 1
            print(f'{current_item_info}: Scraped successfully. SKU: {sku}, Price: {price}, Stock: {stock_availability}')

            # --- Periodic Save and Upload ---
            if items_scrapped > 0 and items_scrapped % 50 == 0:
                print(f'\nSaving sheet after {items_scrapped} items...')
                try:
                    wb.save(file_path)
                    print("Sheet saved locally.")
                    # Attempt FTP upload
                    if not upload_to_ftp(file_path, file_name):
                         print("Periodic FTP upload failed. Continuing script.") # Log failure but continue
                except Exception as save_err:
                    print(f"Error occurred during periodic save/upload: {save_err}")
                    print(traceback.format_exc())

            # Add a small delay to be polite to the server
            time.sleep(2) # 2-second delay between requests

        except Exception as item_error:
            print(f"!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
            print(f"Error processing {current_item_info} for URL {url}: {item_error}")
            print(traceback.format_exc())
            print(f"!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
            # Continue to the next item
            continue

    # --- Final Save and Upload ---
    print("\nScraping loop finished.")
    print(f"Attempting final save for {file_name}...")
    wb.save(file_path)
    print("Final local save successful.")

    # Final FTP upload
    if upload_to_ftp(file_path, file_name):
        print("Final FTP upload successful.")
    else:
        print("Final FTP upload failed.")
        # Consider sending a specific alert or logging this failure more prominently

    # --- Success Notification ---
    print(f"\n{description} scraper completed successfully.")
    print(f"Total items processed in sheet: {item_number}")
    print(f"Items updated/scraped in this run: {items_scrapped}")
    send_email_notification(True, items_scrapped, scraper_name=description)

except Exception as e:
    # --- Error Handling and Notification ---
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    print(f"An unexpected error occurred in the {description} scraper:")
    print(error_message)
    print(f"Traceback:\n{full_traceback}")
    print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")

    # Attempt to save progress if workbook was loaded
    if wb:
        try:
            print("Attempting to save progress before exiting due to error...")
            wb.save(file_path)
            print("Progress saved locally.")
            # Optionally try to upload the partially updated file
            upload_to_ftp(file_path, file_name)
        except Exception as save_error:
            print(f"Could not save progress after error: {save_error}")

    # Send failure email notification
    send_email_notification(False, items_scrapped, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}", scraper_name=description)
    sys.exit(1) # Exit with error code 1 to indicate failure to the GitHub Action runner

finally:
    # Close the workbook if it was opened (optional, as context manager isn't used here)
    # if wb:
    #     wb.close() # openpyxl doesn't strictly require closing like file handles
    print(f"{description} script finished.")
