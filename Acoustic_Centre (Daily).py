import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium_stealth import stealth
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from send2trash import send2trash
from datetime import datetime, timedelta

import pprint, time, math, os

wb = openpyxl.load_workbook(rf"\\SERVER\Python\Pricing\Pricing Spreadsheets\Acoustic_Centre.xlsx")
sheet = wb['Sheet']

url_list = []
item_number = 0

items_scrapped = 0
counter = 0


for x in range(2, sheet.max_row+1):
    url_list.append(sheet['E'+str(x)].value)

r = requests.get(fr'https://www.acousticcentre.com.au/search?page=1&q=+&type=product')

soup = BeautifulSoup(r.content, 'html.parser')

soup = soup.find(class_='dropdown_container mega-menu mega-menu-5')
for sheet_line in range(2, sheet.max_row+1):

    sheet_line-=counter

    further_break = ''

    item_number +=1

    time_last_scrapped = sheet['H' + str(sheet_line)].value
    string_datetime_conversion = datetime.strptime(time_last_scrapped, '%m %d %Y')
    if string_datetime_conversion + timedelta(days=7) > datetime.today():
        print(f'Item {str(item_number)} scrapped less than 7 days ago, skipping...')

    else:
        try:
            url = sheet['E'+str(sheet_line)].value

            r = requests.get(url)
            soup3 = BeautifulSoup(r.content, 'html.parser')

            try:
                sku = soup3.find(class_='sku').text.strip()
            except:
                continue
            image = soup3.find(class_='image__container')
            image = image.find('img')['data-src']
            price = soup3.find(class_='price-ui')
            price = price.find(class_='price').text
            price = price.replace('$', '')
            price = price.replace(',', '')

            try:
                description = soup3.find(id='tabs-2').text.strip()

            except:
                'Description not avaliable.'
            print(description)

            stock = soup3.find(class_='purchase-details').text
            stock = stock.replace('\n', '')
            stock = stock.strip()

            if 'cart' in stock:
                stock_avaliability = 'y'
            else:
                stock_avaliability = 'n'

            today = datetime.now()

            date = today.strftime('%m %d %Y')

            items_scrapped+=1
            item_number+=1
            sheet['A' + str(sheet_line)].value = sku
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliability



            print(f'Item {str(item_number)} scraped successfully')

            if int(items_scrapped) % 20 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(rf"\\SERVER\Python\Pricing\Pricing Spreadsheets\Acoustic_Centre.xlsx")
                except:
                    print(f"Error occurred while saving the Excel file:")
        except:
            sheet.delete_rows(sheet_line, 1)
            counter += 1

try:
    wb.save(rf"\\SERVER\Python\Pricing\Pricing Spreadsheets\Acoustic_Centre.xlsx")
except:
    print(f"Error occurred while saving the Excel file:")





