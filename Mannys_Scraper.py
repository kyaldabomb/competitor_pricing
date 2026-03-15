import requests
import time, os, traceback
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import sys
import argparse
import openpyxl
import ftplib

from typesense_fetch import (
    fetch_all_products, construct_image_url, construct_product_url,
    strip_html, build_url_lookup, build_slug_lookup, TARGET_BRANDS,
)
from supplier_sku import build_all_lookups, resolve_sku
from sku_remapping import apply_sku_remapping

# Parse command line arguments
parser = argparse.ArgumentParser(description='Run Mannys scraper (discovers new products + updates existing)')
parser.add_argument('scraper', nargs='?', default='mannys',
                    help='Scraper name from config')
args = parser.parse_args()


def upload_to_ftp(file_path, file_name):
    print(f"\n==== Uploading {file_name} to FTP ====")
    try:
        ftp_password = os.environ.get('FTP_PASSWORD')
        if not ftp_password:
            print("FTP_PASSWORD not found in environment variables")
            return False

        session = ftplib.FTP('ftp.drivehq.com', 'kyaldabomb', ftp_password)

        if 'competitor_pricing' not in session.nlst():
            print("competitor_pricing directory not found, creating it...")
            session.mkd('competitor_pricing')

        session.cwd('competitor_pricing')

        with open(file_path, 'rb') as file:
            session.storbinary(f'STOR {file_name}', file)

        with open('upload_timestamp.txt', 'w') as f:
            f.write(f"Upload of {file_name} completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        with open('upload_timestamp.txt', 'rb') as file:
            session.storbinary('STOR upload_timestamp.txt', file)

        session.quit()
        print(f"File {file_name} uploaded to FTP successfully")
        return True
    except Exception as e:
        print(f"Error uploading to FTP: {str(e)}")
        print(traceback.format_exc())
        return False


def send_email_notification(success, new_count=0, updated_count=0, error_msg=""):
    print("Sending email notification...")
    try:
        sender = "kyal@scarlettmusic.com.au"
        receiver = "kyal@scarlettmusic.com.au"
        password = os.environ.get('EMAIL_PASSWORD')
        if not password:
            print("Email password not found in environment variables")
            return

        host = "mail.scarlettmusic.com.au"
        port = 587

        msg = MIMEMultipart()
        msg['From'] = sender
        msg['To'] = receiver

        if success:
            msg['Subject'] = f"Mannys Scraper Success: {new_count} new, {updated_count} updated"
            body = (
                f"The Mannys web scraper ran successfully.\n\n"
                f"New products added: {new_count}\n"
                f"Existing products updated: {updated_count}\n"
            )
        else:
            msg['Subject'] = "Mannys Scraper Failed"
            body = f"The Mannys web scraper encountered an error:\n\n{error_msg}"

        msg.attach(MIMEText(body, 'plain'))

        server = smtplib.SMTP(host, port)
        server.starttls()
        server.login(sender, password)
        server.send_message(msg)
        server.quit()
        print("Email notification sent successfully")
    except Exception as e:
        print(f"Failed to send email notification: {str(e)}")
        print(traceback.format_exc())


# Brand filter matching — replicates the original scraper's filter logic
BRAND_FILTERS = [
    'orange', 'ernie ball', 'korg', 'arturia', 'jbl', 'epiphone',
    'dbx', "d'addario", 'tech 21', 'lr baggs', 'universal audio',
    'soundcraft', 'aguilar', 'casio', 'akg', 'se',
    'morely', 'blue microphones', 'soundbrenner', 'strandberg',
]

BRAND_CONTAINS = [
    'tc electronic', 'gibson', 'seymour', 'helicon', 'kyser', 'gruv',
    'akai', 'marshall', 'nord', 'hercules', 'headrush', 'boss',
    'ashton', 'ibanez', 'evans', 'tascam', 'gator', 'valencia',
    'xtreme', 'cnb', 'v-case', 'mahalo', 'dxp', 'dunlop', 'mano',
    'carson', 'mxr', 'armour', 'dimarzio', 'auralex', 'alesis',
    'digitech', 'crown', 'samson', 'x-vive', 'beale', 'snark',
    'esp', 'ghs', 'strymon', 'rockboard', 'vic firth', 'ik multimedia',
    'remo', 'darkglass', 'martin', 'm-audio', 'native instruments',
    'source audio', 'emg', 'mapex', 'udg', 'alto', 'nektar',
    'radial', 'teenage', 'tama', 'roland', 'hosa', 'oskar',
    'hotone', 'vox', 'ampeg', 'singular',
]


def is_target_brand(brand_name):
    """Check if a brand matches the original scraper's filter logic."""
    bl = brand_name.lower()
    if bl in BRAND_FILTERS:
        return True
    for pattern in BRAND_CONTAINS:
        if pattern in bl:
            return True
    return False


try:
    file_path = "Pricing Spreadsheets/Mannys.xlsx"
    file_name = "Mannys.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet']

    # Build set of existing URLs for O(1) lookup
    url_set = set()
    for x in range(2, sheet.max_row+1):
        url = sheet['E'+str(x)].value
        if url:
            url_set.add(url)

    print(f"Found {len(url_set)} existing URLs in the spreadsheet")

    # ================================================================
    # FETCH: Get all products from Mannys Typesense API (one call)
    # ================================================================
    print("\n==== Fetching all products from Mannys API ====")
    all_products = fetch_all_products(brands=TARGET_BRANDS)
    print(f"Fetched {len(all_products)} total products from API")

    # Group products by brand (needed for Phase 1)
    products_by_brand = {}
    for p in all_products:
        brand = p.get('brand', '')
        if brand not in products_by_brand:
            products_by_brand[brand] = []
        products_by_brand[brand].append(p)

    print(f"Products grouped into {len(products_by_brand)} brands")

    # Build supplier SKU lookups for auto-resolution (needed for Phase 1)
    brands_needed = set()
    for brand_name in products_by_brand:
        if is_target_brand(brand_name):
            brands_needed.add(brand_name.lower())

    print("\n==== Building supplier SKU lookups ====")
    supplier_lookups = build_all_lookups(brands_needed)

    # ================================================================
    # PHASE 1: Add new products (monthly logic)
    # ================================================================
    print("\n==== Phase 1: Discovering new products ====")
    new_count = 0
    auto_resolved_count = 0
    manual_mapped_count = 0
    item_number = 0

    for brand_name, brand_products in sorted(products_by_brand.items()):
        if not is_target_brand(brand_name):
            continue

        brand_lower = brand_name.lower()
        supplier_lookup = supplier_lookups.get(brand_lower, {})

        for product in brand_products:
            item_number += 1

            try:
                url = construct_product_url(product.get('url', ''))

                if url in url_set:
                    continue

                sku = product.get('sid', '')
                if not sku:
                    continue

                brand = brand_name

                # Try auto-SKU resolution via supplier feeds first
                auto_sku = resolve_sku(sku, brand_lower, supplier_lookup)

                if auto_sku:
                    sku = auto_sku
                    auto_resolved_count += 1
                else:
                    remapped = apply_sku_remapping(sku, brand)
                    if remapped is None:
                        continue
                    sku = remapped
                    manual_mapped_count += 1

                title = product.get('title', 'N/A')
                price = str(product.get('price', 'N/A'))
                stock_avaliable = 'y' if product.get('stock') in ('In Stock', 'Low Stock') else 'n'

                image = 'N/A'
                if product.get('image'):
                    image = construct_image_url(product['image'][0])

                description = strip_html(product.get('description', ''))
                if not description:
                    description = 'N/A'

                date = datetime.now().strftime('%m %d %Y')

                sheet.append([sku, brand, title, price, url, image, description, date, stock_avaliable])
                url_set.add(url)

                new_count += 1
                if new_count % 50 == 0:
                    print(f'  New products found: {new_count}...')

            except Exception as product_error:
                print(f"Error processing new product {item_number}: {str(product_error)}")
                print(traceback.format_exc())

    print(f"\nPhase 1 complete: {new_count} new products added")
    print(f"  Auto-resolved SKUs: {auto_resolved_count}")
    print(f"  Manual-mapped SKUs: {manual_mapped_count}")

    # Save after Phase 1 if we added new products
    if new_count > 0:
        print("Saving after Phase 1...")
        wb.save(file_path)
        upload_to_ftp(file_path, file_name)

    # ================================================================
    # PHASE 2: Update existing stale products (daily logic)
    # ================================================================
    print("\n==== Phase 2: Updating existing products ====")

    # Build lookup dicts from the same API data
    url_lookup = build_url_lookup(all_products)
    slug_lookup = build_slug_lookup(all_products)

    updated_count = 0

    for sheet_line in range(2, sheet.max_row+1):
        try:
            # Check if stale (>7 days since last scrape)
            time_last_scrapped = sheet['H' + str(sheet_line)].value
            if time_last_scrapped:
                try:
                    last_date = datetime.strptime(time_last_scrapped, '%m %d %Y')
                    if last_date + timedelta(days=7) > datetime.today():
                        continue
                except Exception as date_error:
                    print(f"Error parsing date {time_last_scrapped}: {str(date_error)}")

            url = sheet['E'+str(sheet_line)].value
            if not url:
                continue

            # Match against API data
            product = url_lookup.get(url)

            if not product:
                old_slug = url.rstrip('/').split('/')[-1]
                product = slug_lookup.get(old_slug)
                if product:
                    sheet['E' + str(sheet_line)].value = construct_product_url(product['url'])

            if not product:
                continue

            # Update fields
            sheet['D' + str(sheet_line)].value = str(product['price'])
            sheet['F' + str(sheet_line)].value = construct_image_url(product['image'][0]) if product.get('image') else 'N/A'
            sheet['G' + str(sheet_line)].value = strip_html(product.get('description', ''))
            sheet['H' + str(sheet_line)].value = datetime.now().strftime('%m %d %Y')
            sheet['I' + str(sheet_line)].value = 'y' if product.get('stock') in ('In Stock', 'Low Stock') else 'n'

            updated_count += 1

            if updated_count % 500 == 0:
                print(f'  Updated: {updated_count} products...')
                wb.save(file_path)

        except Exception as item_error:
            print(f"Error updating row {sheet_line}: {str(item_error)}")
            print(traceback.format_exc())

    print(f"\nPhase 2 complete: {updated_count} existing products updated")

    # ================================================================
    # FINAL: Save, upload, notify
    # ================================================================
    wb.save(file_path)
    upload_to_ftp(file_path, file_name)
    print(f"\n==== Scraping completed successfully ====")
    print(f"  New products added: {new_count}")
    print(f"  Existing products updated: {updated_count}")
    send_email_notification(True, new_count=new_count, updated_count=updated_count)

except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")

    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")

    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)
