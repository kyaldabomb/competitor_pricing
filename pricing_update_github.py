import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import exceptions
import csv
import ftplib
import xlrd
from datetime import date
import requests, json, pprint, os, re, io, math
from csv import reader
import math
import glob
from csv import reader
from datetime import datetime, timedelta


def is_date_over_a_year_ago(past_date_str):
    """
    Checks if a given date string is over a year ago compared to today's date.
    """
    try:
        past_date = datetime.strptime(past_date_str, "%d/%m/%Y")
    except ValueError:
        return None  # Handle invalid date format

    today = datetime.now()
    one_year_ago = today - timedelta(days=365)  # Account for leap years later

    return past_date < one_year_ago


def Auralex(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    sku = sku

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Alesis(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    sku = sku
    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.9
    promo_success = 'n'
    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.75)
    except:
        cost = (RRP * 0.75)

    if promo_success == 'y':
        return RRP, cost
    

    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:
            return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Akai(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    sku = sku
    if '69mpkmini3' in sku.lower():

        cost = (RRP * 0.7) * 0.82
    else:
        cost = (RRP * 0.7) * 0.95

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Ashton(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    original_RRP = float(RRP)
    default_discount = 0.85
    promo_success = 'y'


    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'n'

                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.85
    except:
        cost = (RRP * 0.7) * 0.85

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Armour(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)
    default_discount = 1
    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.8
    except:
        cost = (RRP * 0.7) * 0.8

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def AKG(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)
    default_discount = 0.85
    promo_success = 'n'


    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                    default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Arturia(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)
    default_discount = 0.85
    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'n'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.95

    except:
        cost = (RRP * 0.7) * 0.95

    if promo_success == 'y':
        return RRP, cost

    if RRP >= 100:
        return RRP * default_discount, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Aguilar(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)
    default_discount = 0.85
    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.85

    except:
        cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost

    if RRP >= 100:
        return RRP * default_discount, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Beale(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

def Dean(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)*0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*0.9, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Blackstar(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.8

    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:

        if 'power supply' in title.lower():
            cost = RRP * 0.7
        else:

            cost = (RRP * 0.7) * 0.85

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:
            return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Boss(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = RRP * 0.75

    if RRP >= 150:
        return RRP*0.85, cost
    if 25 <= RRP < 150:
        return RRP, cost

    if 0 <= RRP < 25:
        return RRP + 5, cost


def Casio(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                    default_discount = 1
            break

    try:
        cost

    except:
        cost = RRP * 0.7

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def ColeClark(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if RRP >= 50:
        return RRP, cost
    if 25 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 25:
        return RRP + 5, cost


def Behringer(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    sku = sku
    original_RRP = float(RRP)
    default_discount = 1
    promo_success = 'n'
    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                return RRP, cost

        try:
            cost

        except:
            cost = (RRP * 0.7)
    except:
        cost = cost = RRP * 0.7
    

    behringer_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Behringer.xlsx")
    behringer_sheet = behringer_workbook['Sheet1']
    for items in range(1, behringer_sheet.max_row + 1):
        behringer_sku = str(behringer_sheet['D' + str(items)].value)
        if behringer_sku is None:
            continue
        if behringer_sku in sku:
            cost = str(behringer_sheet['I' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            cost = float(cost) * 1.1
            break

    try:
        cost

    except:
        cost = RRP * 0.7
    if promo_success == 'y':
        return RRP, cost
    if RRP >= 35:
        return RRP, cost

    if 0 <= RRP < 35:
        return RRP + 5, cost


def Bluegrass(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = RRP * 0.59

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Cordoba(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if RRP >= 100:
        return RRP * 0.85, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Celestion(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.88

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Carson(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.55)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def CNB(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.59)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Crossfire(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    title = title

    cost = (RRP * 0.7) * 0.7

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Crown(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    title = title

    if 'metronome' in title.lower() or 'piano' in title.lower() or 'keyboard' in title.lower() or 'bench' in title.lower() or 'seat' in title.lower():

        cost = (RRP * 0.7) * 0.7

    else:
        cost = (RRP * 0.7) * 0.85

    if RRP >= 100:
        return RRP * 0.85, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Daddario(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    title = title
    sku = sku
    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.7

    if '-3D' in sku:
        cost = (RRP * 0.7) * 0.8
    if '-10P' in sku:
        cost = (RRP * 0.7) * 0.9
    if '-B25' in sku:
        cost = (RRP * 0.7) * 0.9

    if RRP >= 50:
        return RRP, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 2, cost


def Darkglass(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*0.78, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Tasman(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)


    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.8

    if RRP >= 100:

        return RRP * 0.85, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def DXP(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.63)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Dunlop(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    if 'tortex' in title.lower():
        cost = (RRP * 0.51)
    else:
        cost = RRP * 0.59

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def ESP(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.8
    promo_success = 'n'
    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:

        if "e2" in sku.lower():
            cost = (RRP * 0.7)

        else:
            cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == "Y":
        if RRP >= 100:
            return cost * 1.05, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 100:
            return RRP * default_discount, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def JimDunlop(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    if 'tortex' in title.lower():
        cost = (RRP * 0.51)
    else:
        cost = (RRP * 0.59)

    if RRP >= 100:
        return RRP * 0.85, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Dimarzio(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.59)

    if RRP >= 100:
        return RRP * 0.85, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def DBX(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.85
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost
    if RRP >= 100:
        return RRP * default_discount, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Digitech(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.85
    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.85

    except:
        cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == "Y":
        if RRP >= 100:
            return cost * 1.05, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 100:
            return RRP * default_discount, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Ebow(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if RRP >= 100:
        return RRP * 0.85, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def ErnieBall(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Ernie_Ball.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['E' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['D' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            break

    try:
        cost

    except:

        if 'bass' in title.lower() and 'str' in title.lower():
            cost = (RRP * 0.7) * 0.8

        elif 'music man' in title.lower():
            cost = (RRP * 0.7)


        elif 'strap' in title.lower():
            cost = (RRP * 0.7) * 0.7

        elif 'cable' in title.lower() or 'slide' in title.lower():
            cost = (RRP * 0.7) * 0.7

        elif 'paradigm' in title.lower():
            cost = (RRP * 0.7) * 0.85

        elif 'm-steel' in title.lower():
            cost = (RRP * 0.7) * 0.8


        else:

            cost = (RRP * 0.7) * 0.6

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def MusicMan(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.8

    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Ernie_Ball.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['E' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['D' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7)

    # cost = (RRP * 0.7)
    if promo_success == 'y':
        return RRP, cost
        
    if RRP >= 100:
        return RRP * default_discount, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Sterling(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.85
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Ernie_Ball.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['E' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['D' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.85

    # cost = (RRP * 0.7)*0.85
    if promo_success == 'y':
        return RRP, cost
    if RRP >= 50:
        return RRP * default_discount, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Epiphone(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    # cost = (RRP*0.7)*0.9
    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.8
    #######################################
    promo_success = 'n'
    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.9
    except:
        cost = (RRP * 0.7) * 0.9

    if promo_success == 'y':
        return RRP, cost

    #########################################################
    if obsolete_stock == "Y":
        if RRP >= 100:
            return cost * 1.05, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 100:
            return RRP * default_discount, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Evans(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.63)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Gator(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)
    default_discount = 1

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.85
    except:
        cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost


    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Gibson(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.85

    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.9
    except:
        cost = (RRP * 0.7) * 0.9

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == "Y":
        if RRP >= 100:
            return cost * 1.05, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 100:
            return RRP * default_discount, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Headrush(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.95

    if RRP >= 100:
        return RRP * 0.85, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Hercules(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.7

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Ibanez(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.80

    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1

                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.8
    except:
        cost = (RRP * 0.7) * 0.8

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == "Y":
        if RRP >= 100:
            return cost * 1.05, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 100:
            return RRP * default_discount, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Jargar(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.6)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def JBL(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    sku = sku

    RRP = float(RRP)
    default_discount = 1
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                    default_discount = 1
            break

    try:
        cost

    except:

        cost = (RRP * 0.8)

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def GruvGear(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.8

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def GHS(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    default_discount = 1
    original_RRP = float(RRP)

    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.8
    except:
        cost = (RRP * 0.7) * 0.8

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def IKMultimedia(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.75)

    if RRP >= 50:
        return RRP * 0.8, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Kaces(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Kawai(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.98

    if RRP >= 50:
        return RRP * 0.85, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Kyser(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.8

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Korg(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    title = title

    RRP = float(RRP)
    default_discount = 0.9
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            cost = float(cost) * 1.1
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                    default_discount = 1
            break

    try:
        cost

    except:
        if 'tuner' in title.lower():

            if 'orchest' in title.lower():
                cost = (RRP * 0.7) * 0.9
            else:

                cost = (RRP * 0.7) * 0.65
        else:
            cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost
    if RRP >= 100:
        return RRP * default_discount, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def LRBaggs(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)
    default_discount = 0.8
    promo_success = 'n'
    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.8
    except:
            cost = (RRP * 0.7) * 0.8
    
    if promo_success == 'y':
        return RRP, cost
    if RRP >= 50:
        return RRP * default_discount, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Mahalo(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.63)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Marshall(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    


    obsolete_stock = obsolete_stock
    title = title
    default_discont = 0.85

    if 'fridge' in title.lower():
        cost = (RRP * 0.75)
        default_discont = 0.9

    else:

        cost = (RRP * 0.7) * 0.8

    if RRP >= 100:
        return RRP * default_discont, cost
    if 10 <= RRP < 100:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Mano(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.55)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Martinez(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            cost = float(cost) * 1.1
            promo_success = 'y'
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.7
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def MartinStrings(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.7

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def MAudio(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Monterey(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            cost = float(cost)
            promo_success = 'y'
            break

    try:
        cost
    except:

        cost = (RRP * 0.7) * 0.75

    if promo_success == 'y':
            return RRP, cost
    
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Mitello(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.65)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Mooer(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.7

    if RRP >= 50:
        return RRP * 0.8, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def MXR(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.59)

    if RRP >= 50:
        return RRP * 0.8, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Laney(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.59)

    if RRP >= 50:
        return RRP * 0.8, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def NativeInstruments(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = RRP * 0.7

    if RRP >= 50:
        return RRP * 0.85, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Nektar(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = RRP * 0.7

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def SE(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = RRP * 0.7

    if RRP >= 50:
        return RRP * 0.9, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Nord(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = RRP * 0.7

    if RRP >= 50:
        return RRP * 0.87, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def PlanetWaves(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.8

    if RRP >= 50:
        return RRP * 0.9, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Orange(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    default_discount = 0.9
    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            if 'crush' in title.lower():
                cost = (RRP * 0.7) * 0.75

            if 'uk' in title.lower() and 'pedal' not in title.lower():
                cost = (RRP * 0.7)
            else:
                cost = (RRP * 0.7) * 0.85
    except:
        if 'crush' in title.lower():
            cost = (RRP * 0.7) * 0.75

        if 'uk' in title.lower() and 'pedal' not in title.lower():
            cost = (RRP * 0.7)
        else:
            cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP* default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Paytons(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 1
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7)*0.9

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Pirastro(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.68)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Promark(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.56)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Powerbeat(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.69)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Radial(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Rapco(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Rico(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.56)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Remo(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    title = title

    if "head" in title.lower():

        cost = (RRP * 0.49)

    else:
        cost = (RRP * 0.55)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Rockboard(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Sanchez(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.7

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Samson(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Schaller(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.6)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Steinhoff(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.7

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Snark(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.59)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Strauss(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.7

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def SeymourDuncan(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.85
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7)
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:
            return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Tanglewood(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)
    promo_success = 'n'

    try:

        default_discount = 0.8
    
        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)
    
                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break
    
        try:
            cost
    
        except:
            cost = (RRP * 0.7) * 0.8
    except Exception as e:
        print(e)
    if promo_success == 'y':
        return RRP, cost
    if RRP >= 200:
        return RRP * 0.8, cost
    if 80 <= float(RRP) < 200:
        return RRP * 0.9, cost
    if 25 <= RRP < 80:
        return RRP, cost
    if 0 <= RRP < 25:
        return RRP + 5, cost


def Tascam(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.8

    sku = sku
    promo_success = 'n'
    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:

            if "DA-6400DP" in sku:

                 return RRP, cost

            else:

                return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def TCElectronic(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 1
    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.75)
    except:
        cost = (RRP * 0.75)

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:
            return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def TCHelicon(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.9

    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = cost = (RRP * 0.75)
    except:
        cost = (RRP * 0.75)

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:
            return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def TG(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if RRP >= 50:
        return RRP * 0.9, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Tech21(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.85
    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.9
    except:
        cost = (RRP * 0.7) * 0.9

    cost = (RRP * 0.7) * 0.9

    if promo_success == 'y':
        return RRP, cost

    if RRP >= 50:
        return RRP * 0.8, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def UniversalAudio(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

  

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.8
    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)
    
                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                promo_success = 'y'
                if original_RRP * default_discount > float(RRP):
                    default_discount = 1
                break
    
        try:
            cost
    
        except:
            cost = (RRP * 0.7) * 0.85

    except Exception as e:

        print(e)

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:
            return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def UltimateSupport(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.6

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Savarez(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.8

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Sequenz(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def SourceAudio(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.8

    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.85

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:
            return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Spector(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock
    promo_success = 'n'

    default_discount = 0.8
    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.85

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:
            return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Soundcraft(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    promo_success = 'n'

    default_discount = 0.8
    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == "Y":

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:
        if RRP >= 50:
            return RRP * default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def SonicDrive(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Strymon(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = RRP * 0.7

    if RRP >= 50:
        return RRP * 0.85, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Tourtek(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.7

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Valencia(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.63)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def VicFirth(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Xtreme(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.59)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def XTR(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.55)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Xvive(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if RRP >= 50:
        return RRP * 0.9, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def VCase(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.59)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def UXL(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.7
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def BarnesMullins(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    default_discount = 0.8
    original_RRP = float(RRP)
    promo_success = 'n'


    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.8
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def EMG(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Hardcase(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Toca(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.95

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Gibraltar(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Onstage(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def MusicNomad(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Alto(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.8

    if RRP >= 50:
        return RRP * 0.85, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Essex(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.63)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Boomwhackers(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def NUX(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Lanikai(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def TeenageEngineering(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Shubb(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Progressive(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if RRP >= 50:
        return RRP + 4, cost
    if 10 <= RRP < 50:
        return RRP + 4, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Shure(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Alvarez(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Tama(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Hartke(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if RRP >= 50:
        return RRP * 0.85, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Basso(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def LeeOskar(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.8

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Aquila(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Roland(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.75)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 150:
            return RRP*0.85, cost
        if 10 <= RRP < 150:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Hotone(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Hosa(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.72

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Duracell(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            break

    try:
        cost

    except:
        cost = (RRP * 0.7)
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Grover(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.55)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def DCM(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)


    default_discount = 1
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.8

    
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Hidersine(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)
    default_discount = 0.85
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.9
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Cioks(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = RRP * 0.7

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Puresound(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.78

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def JohnPearse(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.95

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def KandK(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Hamilton(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Alpine(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.7
    except:
        cost = (RRP * 0.7) * 0.7
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Aquarian(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    promo_success = 'n'

    try:

        gibson_workbook = openpyxl.load_workbook(
            rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
        gibson_sheet = gibson_workbook['Sheet1']
        for items in range(1, gibson_sheet.max_row + 1):
            gibson_sku = str(gibson_sheet['A' + str(items)].value)
            if gibson_sku is None:
                continue
            if gibson_sku.lower() == sku.lower():
                RRP = str(gibson_sheet['L' + str(items)].value)
                RRP = RRP.replace('$', '')
                RRP = RRP.replace(',', '')
                RRP = float(RRP)

                cost = str(gibson_sheet['F' + str(items)].value)
                cost = cost.replace('$', '')
                cost = cost.replace(',', '')
                cost = float(cost)
                promo_success = 'y'
                break

        try:
            cost

        except:
            cost = (RRP * 0.7) * 0.6
    except:
        cost = (RRP * 0.7) * 0.6

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Maestro(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Modal(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.9

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Netgear(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Onestage(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.8)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Lexicon(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.85

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Vox(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)

    obsolete_stock = obsolete_stock

    default_discount = 0.85

    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == "Y":
        if RRP >= 100:
            return cost * 1.05, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 100:
            return RRP * default_discount, cost
        if 10 <= RRP < 100:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Hagstrom(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Vater(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Markbass(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.8

    if RRP >= 50:
        return RRP * 0.85, cost
    if 10 <= RRP < 50:
        return RRP, cost

    if 0 <= RRP < 10:
        return RRP + 5, cost


def Deering(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Lag(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)
    promo_success = 'n'
    obsolete_stock = obsolete_stock

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.85
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Avid(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.85)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Guild(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7) * 0.9

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Eikon(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    original_RRP = float(RRP)
    default_discount = 0.8
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            promo_success = 'y'
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.8

    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def National(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)
    original_RRP = float(RRP)
    obsolete_stock = obsolete_stock

    default_discount = 0.85
    promo_success = 'n'
    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            if original_RRP * default_discount > float(RRP):
                default_discount = 1

            promo_success = 'y'
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.9

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

def Boveda(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost


def Elixir(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)
    original_RRP = float(RRP)
    obsolete_stock = obsolete_stock
    default_discount = 0.7
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            promo_success = 'y'
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.7

    
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*default_discount, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

def Rotosound(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)
    original_RRP = float(RRP)
    obsolete_stock = obsolete_stock

    default_discount = 0.7
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            promo_success = 'y'
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.7

    if promo_success == 'y':
        return RRP, cost

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*0.7, cost
        if 10 <= RRP < 50:
            return RRP*0.8, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

def Kink(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5
    original_RRP = float(RRP)
    RRP = float(RRP)

    obsolete_stock = obsolete_stock
    default_discount = 0.83
    promo_success = 'n'

    gibson_workbook = openpyxl.load_workbook(
        rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx")
    gibson_sheet = gibson_workbook['Sheet1']
    for items in range(1, gibson_sheet.max_row + 1):
        gibson_sku = str(gibson_sheet['A' + str(items)].value)
        if gibson_sku is None:
            continue
        if gibson_sku.lower() == sku.lower():
            RRP = str(gibson_sheet['L' + str(items)].value)
            RRP = RRP.replace('$', '')
            RRP = RRP.replace(',', '')
            RRP = float(RRP)

            cost = str(gibson_sheet['F' + str(items)].value)
            cost = cost.replace('$', '')
            cost = cost.replace(',', '')
            if original_RRP * default_discount > float(RRP):
                default_discount = 1
            promo_success = 'y'
            break

    try:
        cost

    except:
        cost = (RRP * 0.7) * 0.9

    
    if promo_success == 'y':
        return RRP, cost
    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*0.83, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

def SwiffAudio(RRP, title, sku, obsolete_stock):
    # High end (>200) = *0.7
    # mid range (80-200) = *0.8
    # low range (25-80) = *1
    # super low range (0-25) = +5

    RRP = float(RRP)

    obsolete_stock = obsolete_stock

    obsolete_stock = obsolete_stock

    cost = (RRP * 0.7)

    if obsolete_stock == 'Y':

        if RRP >= 50:
            return cost * 1.05, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

    else:

        if RRP >= 50:
            return RRP*0.83, cost
        if 10 <= RRP < 50:
            return RRP, cost

        if 0 <= RRP < 10:
            return RRP + 5, cost

#Ibanez removed, pricing doesn't update
#'esp': ESP, 'esp guitars': ESP, 'esp e-2': ESP,'esp ltd': ESP,'espltd': ESP, 'korg': Korg, 'vox': Vox, "universal audio": UniversalAudio, "ua": UniversalAudio, "uaguitar": UniversalAudio , removed, 

completed_brands = {"tanglewood": Tanglewood,"dean": Dean, "morley": Dean, "orange": Orange, 'ernie ball': ErnieBall,
                    'arturia': Arturia, 'strandberg': Arturia, 'jbl': JBL, 'epiphone': Epiphone, 'gibson': Gibson,
                    'tc electronic': TCElectronic, 'dbx': DBX, "daddario": Daddario, "planet waves": PlanetWaves,
                    "tech 21": Tech21, "lr baggs": LRBaggs, "aguilar": Aguilar,  "soundcraft": Soundcraft, "avid": Avid, "behringer": Behringer, "casio": Casio, "akg": AKG, "drstrings": AKG,
                    "blackstar": Blackstar,
                    "tc helicon": TCHelicon, "seymour duncan": SeymourDuncan, "seymour": SeymourDuncan,
                    "gruv gear": GruvGear, "kyser": Kyser,
                    "akai": Akai, 'marshall': Marshall, 'nord': Nord, 'hercules': Hercules, 'celestion': Celestion
    , 'boss': Boss, 'general products': Boss, 'ashton': Ashton, 'headrush': Headrush, 'cordoba': Cordoba, 'evans': Evans
    , 'promark': Promark, 'rico': Rico, 'gator cases': Gator, 'paytons': Paytons, 'admira': Paytons, 'hills': Paytons,'beam accessories': Paytons,'carl martin': Paytons,'ecostrap': Paytons,'limousine': Paytons, 'tascam': Tascam,
                    'valencia': Valencia, 'xtreme': Xtreme, 'cnb': CNB, 'v-case': VCase, 'mahalo': Mahalo,
                    '1880 ukulele co': Mahalo, 'dxp': DXP
    , 'dunlop': Dunlop, 'jim dunlop': Dunlop, 'xtr': XTR, 'mano percussion': Mano, 'mano': Mano, 'bluegrass': Bluegrass,
                    'carson': Carson, 'carson cable co': Carson, 'mxr': MXR, 'way huge': MXR,
                    'armour': Armour, 'dimarzio': Dimarzio, 'auralex': Auralex, 'alesis': Alesis, 'digitech': Digitech,
                    'cmi': Ebow,
                    'sanchez': Sanchez, 'steinhoff': Steinhoff, 'k steinhoff': Steinhoff, 'crown': Crown,
                    'crossfire': Crossfire, 'strauss': Strauss, 'martinez': Martinez,
                    'mooer': Mooer, 'samson': Samson, 'samson audio': Samson, 'samson wireless': Samson,
                    'mitello': Mitello, 'powerbeat': Powerbeat,
                    'schaller': Schaller, 'jargar': Jargar, 'pirastro': Pirastro, 'sonic drive': SonicDrive,
                    'xvive': Xvive,
                    'beale': Beale, 'snark': Snark, 'kaces': Kaces, 
                    'ghs': GHS, 'strymon': Strymon, 'warwick rockboard': Rockboard, 'sterling': Sterling,
                    'sterling by mm': Sterling
    , 'sterling by mus': Sterling, 'sterling by music man': Sterling, 'ernie ball music man': MusicMan,
                    'ernie ball by music man': MusicMan,
                    'music man': MusicMan, 'vic firth': VicFirth, 'ik multimedia': IKMultimedia, 'remo': Remo, 'tg': TG,
                    'pickboy': TG,
                    'fps': TG, 'clarendon': TG, 'johnson': TG, 'teller': TG, 'kun': TG, 'hq': TG, 'kawai': Kawai,
                    'darkglass': Darkglass,
                    'spector': Spector, 'martin strings': MartinStrings, 'm-audio': MAudio, 'monterey': Monterey,
                    'native instrume': NativeInstruments,
                    'native instruments': NativeInstruments, 'ni': NativeInstruments, 'rapco': Rapco, 'rat': Rapco,
                    'sequenz': Sequenz
    , 'source audio': SourceAudio, 'sourceaud': SourceAudio, 'tourtek': Tourtek, 'uxl': UXL,
                    'barnes & mullins': BarnesMullins,
                    'emg': EMG, 'hardcase': Hardcase, 'mapex': Hardcase, 'udg': Hardcase, 'toca': Toca,
                    'toca percussion': Toca, 'gibraltar': Gibraltar, 'onstage': Onstage, 'onstage desk': Onstage
    , 'on stage': Onstage, 'music nomad': MusicNomad, 'alto professional': Alto, 'alto': Alto,
                    'boomwhackers': Boomwhackers, 'nux': NUX, 'savarez': Savarez, 'nektar': Nektar,
                    'se electronics': SE, 'cole clark': ColeClark, 'maton': ColeClark, 'radial': Radial,
                    'radial tonebone': Radial, 'sx': Essex, 'essex': Essex,
                    'carry on': Blackstar, 'teenage enginee': TeenageEngineering,
                    'teenage engineering': TeenageEngineering, 'shubb': Shubb,
                    'learn to play books': Shubb, 'koala': Shubb, 'progressive': Progressive, 'alvarez': Alvarez,
                    'lp': Alvarez, 'tama': Tama,
                    'loog': Hartke, 'hartke': Hartke, 'lanikai': Lanikai, 'basso': Basso, 'lee oskar': LeeOskar,
                    'aquila': Aquila, 'roland': Roland, 'v-drum': Roland,
                    'Hotone': Hotone, 'hosa technology': Hosa, 'hosa': Hosa, 'duracell': Duracell,
                    'ultimate suppor': UltimateSupport, 'artec': Shubb,
                    'augustine': Shubb, 'aer': Shubb, 'aranjuez': Shubb, 'aria': Shubb,
                    'axl / longpai': Shubb, 'axl/longpai': Shubb, 'blessing': Shubb, 'blitz': Shubb,
                    'toontrack': Duracell, 'skb': Shubb, 'isp technologie': Shubb,
                    'hiscox': Shubb, 'gilman': Hotone, 'gotoh': Shubb, 'carlsbro': Shubb, 'art': Duracell, 'cbi': Shubb,
                    'cherub': Shubb, 'gt': Shubb
    , 'electro-voice': Shubb, 'ernst keller': Shubb, 'pro': Shubb, 'george dennis': Shubb, 'groove juice': Shubb,
                    'heritage': Shubb, 'hohner': Shubb,
                    'la voz': Shubb, 'intelli': Shubb, 'd-grip': Shubb, 'jj electronics': Shubb, 'j michael': Shubb,
                    'j. reynolds': Shubb, 'kat percussion': Shubb
    , 'kealoha': Shubb, 'kohala': Shubb, 'learn to play b': Shubb, 'leem': Shubb, 'leolani': Shubb, 'luthiers': Shubb,
                    'macdaddy': Shubb
    , 'mbt': Shubb, 'mi si': Shubb, 'nady': Shubb, 'nashua': Shubb, 'odessa': Shubb, 'olympia': Shubb,
                    'pro music': Shubb, 'torque': Shubb
    , 'opus': Shubb, 'outlaw': Shubb, 'peace': Shubb, 'percussion plus': Shubb, 'perris': Shubb,
                    'protection rack': Shubb, 'rms': Shubb
    , 'rock tips': Shubb, 'rockfield': Shubb, 'sit': Shubb, 'superslick usa': Shubb, 'superslick': Shubb,
                    'vibes': Shubb, 'vorson': Shubb, 'wilkinson': Shubb, 'wincent': Shubb
    , 'wittner': Shubb, 'wolf': Shubb, 'fretz': Shubb, 'big band': Shubb, 'casino': Shubb, 'drumfire': Shubb,
                    'handy patch': Shubb, 'j&d': Shubb, 'j&d luthiers': Shubb
    , 'kahzan': Shubb, 'mojo': Shubb, 'neowood': Shubb, 'rare audio': Shubb, 'slam': Shubb, 'soundart': Shubb,
                    'sound art': Shubb, 'tiki': Shubb, 'timberidge': Shubb,
                    'tj wilco': Shubb, 'xhl': Shubb, 'grover': Grover, 'laney': Laney, 'dcm': DCM,
                    'hidersine': Hidersine, 'cioks': Cioks, 'shure': Shure, 'puresound': Puresound,
                    'john pearse': JohnPearse, 'k & k': KandK, 'hamilton': Hamilton, 'midas': Alesis, 'tasman': Alesis,
                    'alpine': Alpine, 'smart': Alpine, 'aquarian': Aquarian, 'enki': Blackstar,
                    'maestro': Maestro, 'modal': Modal, 'netgear': Netgear, 'onestage': Onestage, 'lexicon': Lexicon,
                     'hagstrom': Hagstrom, 'vater percussio': Vater, 'vater': Vater,
                    'markbass': Markbass, 'deering': Deering, 'lag': Lag, 'guild': Guild, 'eikon': Eikon,
                    'italian stage': Eikon, 'vivo': National, 'vivo bows': National,
                    'kremona': National, 'krem': National, 'toms line': National, 'aroma': National,
                    'champion': National, 'thomastik': National, 'dr thomastik': National
    , 'thomastik infel': National, 'elixir': Elixir, 'franklin straps': National, 'other brands': National,
                    'g7': National, 'galux': National, 'grassi': National
    , 'ida maria grassi': National, 'w.e. hill & sons': National, 'dycem': National, 'kirlin': National,
                    'kna': National, 'kna pickups': National,
                    'luna': National, 'ogb': National, 'promuco': National, 'pratley': Kink, 'rotosound': Rotosound,
                    'ampeg': National, 'boveda': Boveda, 'kink': Kink, 'northstar': Kink, 'levinson': Kink, 'swiff audio': SwiffAudio, 'ams': Valencia, 'angel': Valencia, 'applecreek': Valencia, 'australasian': Valencia,
                    # Added from AMS_Brand_LIST.csv
                    '1st note': Valencia, 'amsalfred': Valencia, 'amshal leonard': Valencia, 'amsleem': Valencia,
                    'amspromark': Valencia, 'amsrotosound': Valencia, 'amsschott': Valencia, 'amx': Valencia,
                    'arcadia': Valencia, 'atomic': Valencia, 'bandstand': Valencia, 'bandstandlp': Valencia,
                    'bausch': Valencia, 'beadbrain': Valencia, 'bill russell': Valencia, 'brixton': Valencia,
                    'bryden': Valencia, 'clip it': Valencia, 'cpk': Valencia, 'ctb': Valencia, 'ctk': Valencia,
                    'cub': Valencia, 'custom eagle': Valencia, 'dadi': Valencia, 'dan electro': Valencia,
                    'danelectro': Valencia, 'danoelectric': Valencia, 'deluke': Valencia, 'denio': Valencia,
                    'dr parts': Valencia, 'dr partsalfred': Valencia, 'dr. parts': Valencia, 'dresden': Valencia,
                    'd-tronic': Valencia, 'eagle': Valencia, 'eaglealfredalfred': Valencia, 'energizer': Valencia,
                    'esperanto': Valencia, 'essexkjos': Valencia, 'eveready': Valencia, 'evh': Valencia,
                    'extreme': Valencia, 'first note': Valencia, 'fontaine': Valencia, 'foundry': Valencia,
                    'galli': Valencia, 'glockenspiel': Valencia, 'graph tech': Valencia, 'greg bennett': Valencia,
                    'groovetech': Valencia, 'guarneri': Valencia, 'handy': Valencia, 'hearos': Valencia,
                    'helin': Valencia, 'hemingway': Valencia, 'heos': Valencia, 'herco': Valencia,
                    'herdim': Valencia, 'hogalo': Valencia, 'hot line': Valencia, 'iec': Valencia,
                    'jim dunlopthomastik': Valencia, 'kwik fret': Valencia, 'kwikfret': Valencia, 'laney\\t': Valencia,
                    'leemleem': Valencia, 'linnd lunna': Valencia, 'lm': Valencia, 'lm products': Valencia,
                    'lmp': Valencia, 'lup-x': Valencia, 'mahaloroland': Valencia, 'major': Valencia,
                    'mannys': Valencia, 'mano percussionmaton': Valencia, 'mp': Valencia, 'mrx': Valencia,
                    'nichols': Valencia, 'opticare': Valencia, 'ozark': Valencia, 'p&h': Valencia,
                    'paganini': Valencia, 'partsland': Valencia, 'pearl': Valencia, 'platinum': Valencia,
                    'playmaster': Valencia, 'polys': Valencia, 'razz': Valencia, 'redding': Valencia,
                    'rev willy': Valencia, 'rhythm tech': Valencia, 'rocklines': Valencia, 'roswell': Valencia,
                    'samick guitar w': Valencia, 'scottys': Valencia, 'sg style': Valencia, 'sgw': Valencia,
                    'shadow': Valencia, 'slider': Valencia, 'snoopy': Valencia, 'soodlums': Valencia,
                    'stenor': Valencia, 'stentor': Valencia, 'sterisol': Valencia, 'stu box': Valencia,
                    'stuart box': Valencia, 'studio plus': Valencia, 'sx lapsteel': Valencia, 'taktell': Valencia,
                    'teka': Valencia, 'tonica': Valencia, 'tortex': Valencia, 'total': Valencia,
                    'total percussio': Valencia, 'total percussion': Valencia, 'tourte': Valencia, 'trophy': Valencia,
                    'v case': Valencia, 'v parts': Valencia, 'valenciaalfred': Valencia, 'versa': Valencia,
                    'vitoos': Valencia, 'v-parts': Valencia, 'waltons': Valencia, 'wsc': Valencia, 'wylde': Valencia,
                    'x guard': Valencia, 'xrt': Valencia, 'zakk wylde': Valencia
}

print('Loading Inventory worksheet...')

os.makedirs('Pricing Spreadsheets', exist_ok=True)
os.makedirs('Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers', exist_ok=True)
os.makedirs('Pricing Spreadsheets/Master_Pricing_Speadsheet', exist_ok=True)


def download_pricing_files():
    print("Downloading required pricing files from FTP...")
    ftp_host = 'ftp.drivehq.com'
    ftp_user = 'kyaldabomb'
    ftp_password = os.environ.get('FTP_PASSWORD')

    try:
        session = ftplib.FTP(ftp_host, ftp_user, ftp_password)

        # Download inventory.xlsx
        with open('inventory.xlsx', 'wb') as f:
            session.retrbinary('RETR inventory.xlsx', f.write)

        # Download Promotional_Prices.xlsx
        os.makedirs('Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers', exist_ok=True)
        with open('Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx', 'wb') as f:
            session.retrbinary(
                'RETR competitor_pricing/Pricing_spreadsheets_supplied_by_suppliers/Promotional_Prices.xlsx', f.write)
        # Download Marshall.xlsx
        
        os.makedirs('Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers', exist_ok=True)
        with open('Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Marshall.xlsx', 'wb') as f:
            session.retrbinary(
                'RETR competitor_pricing/Pricing_spreadsheets_supplied_by_suppliers/Marshall.xlsx', f.write)

        # Download Ernie_Ball.xlsx
        with open('Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Ernie_Ball.xlsx', 'wb') as f:
            session.retrbinary('RETR competitor_pricing/Pricing_spreadsheets_supplied_by_suppliers/Ernie_Ball.xlsx',
                               f.write)

        # Download Behringer.xlsx
        with open('Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Behringer.xlsx', 'wb') as f:
            session.retrbinary('RETR competitor_pricing/Pricing_spreadsheets_supplied_by_suppliers/Behringer.xlsx',
                               f.write)

        # Download EbayInventoryUpload.csv
        with open('EbayInventoryUpload.csv', 'wb') as f:
            session.retrbinary('RETR EbayInventoryUpload.csv', f.write)

        # Download all competitor pricing files
        files = []
        session.cwd('competitor_pricing')
        session.retrlines('NLST *.xlsx', files.append)

        for file in files:
            if file.endswith('.xlsx') and not file.startswith('Pricing_spreadsheets_supplied_by_suppliers/'):
                print(f"Downloading {file}...")
                with open(f'Pricing Spreadsheets/{file}', 'wb') as f:
                    session.retrbinary(f'RETR {file}', f.write)

        session.quit()
        print("All files downloaded successfully")
    except Exception as e:
        print(f"Error downloading files: {str(e)}")
        raise


download_pricing_files()

wb = openpyxl.load_workbook('inventory.xlsx')

sheet = wb['Sheet']

values = []
duplicate_values = []


final_workbook = openpyxl.Workbook()
final_worksheet = final_workbook.active

final_worksheet['A1'] = 'SKU'
final_worksheet['B1'] = 'Brand'
final_worksheet['C1'] = 'Title'
final_worksheet['D1'] = 'RRP'
final_worksheet['E1'] = 'Price (A)'
final_worksheet['F1'] = 'Price (B)'
final_worksheet['G1'] = 'Price (C)'
final_worksheet['H1'] = 'Price (D)'
final_worksheet['I1'] = 'Cost'
final_worksheet['J1'] = 'Margin'
final_worksheet['K1'] = 'Postage Type'
final_worksheet['L1'] = 'Price (E)'  ## 12% Cofund eBay price
final_worksheet['M1'] = 'Obsolete'
final_worksheet['N1'] = 'Category'

postage_type_dictionary = []

with open('EbayInventoryUpload.csv', 'r', encoding='utf-8') as read_obj:
    # pass the file object to reader() to get the reader object
    csv_reader = reader(read_obj)
    # Iterate over each row in the csv using reader object
    for row in csv_reader:
        sku = row[1]
        postage_type = row[10]

        if postage_type is None or postage_type == '':
            continue

        postage_type_dictionary.append({'sku': f'{sku}', 'postage': f'{postage_type}'})

print('Deleting duplicate inventory skus...')

for i in range(2, sheet.max_row + 1):
    # print(duplicate_values)
    if i % 1000 == 0:
        print(i)
    if sheet.cell(row=i, column=1).value in values:
        duplicate_values.append(sheet.cell(row=i, column=1).value)

        pass  # if already in list do nothing
    else:
        values.append(sheet.cell(row=i, column=1).value)

# pprint.pprint(duplicate_values)
lastsku = sheet['A' + str(sheet.max_row)].value

print('Combining all scrapped worksheets')

all_files = glob.glob("Pricing Spreadsheets/*.xlsx")
scrapped_items = []
for y in all_files:
    print(y)
    if '~$' in y:
        continue

    wb2 = openpyxl.load_workbook(y)
    sheet2 = wb2['Sheet']

    for xx in range(2, sheet2.max_row + 1):
        skip = 'false'
        match = 'false'
        sheet_sku = str(sheet2['A' + str(xx)].value)

        sheet_brand = str(sheet2['B' + str(xx)].value)
        sheet_title = str(sheet2['C' + str(xx)].value)

        # if sheet_sku == 'SN-ZUMA':
        #     pass
        # else:
        #     continue

        # if '8000206' in sheet_sku:
        #     pass
        # else:
        #     continue

        ####vvvvvv DEBUG
        # if sheet_brand is None:
        #     continue
        # if sheet_brand.lower() == 'casio':
        #     pass
        # else:
        #     continue
        ####^^^^^^^^^^ DEBUG

        if sheet_brand is None:
            continue
        if 'arturia' in sheet_brand.lower():
            if 'CMI' not in sheet_sku:
                sheet_sku = f'{sheet_sku}CMI'

        if sheet_brand.lower() == "d'addario":
            sheet_brand = "Daddario"
        if sheet_brand.lower() == "akai professional":
            sheet_brand = "akai"

        if 'orange' in sheet_brand.lower() or 'epiphone' in sheet_brand.lower() or 'gibson' in sheet_brand.lower() or 'tc electronic' in sheet_brand.lower() \
                or 'seymour' in sheet_brand.lower() or 'ibanez' in sheet_brand.lower() or 'gator' in sheet_brand.lower() or 'ashton' in sheet_brand.lower() \
                or 'xvive' in sheet_brand.lower() or 'beale' in sheet_brand.lower() or 'ghs' in sheet_brand.lower() or 'behringer' in sheet_brand.lower() \
                or 'helicon' in sheet_brand.lower() or 'midas' in sheet_brand.lower() or 'alpine' in sheet_brand.lower() or 'aquarian' in sheet_brand.lower() \
                or 'maestro' in sheet_brand.lower() or 'tasman' in sheet_brand.lower():
            if 'AUSTRALIS' not in sheet_sku:
                sheet_sku = f'{sheet_sku}AUSTRALIS'
        if 'jbl' in sheet_brand.lower() or 'dbx' in sheet_brand.lower() or 'universal audio' in sheet_brand.lower() \
                or 'dbx' in sheet_brand.lower() or 'soundcraft' in sheet_brand.lower() or 'aguilar' in sheet_brand.lower() or 'tascam' in sheet_brand.lower() \
                or 'akg' in sheet_brand.lower() or 'digitech' in sheet_brand.lower() or 'cmi' in sheet_brand.lower() or 'esp' in sheet_brand.lower() or 'ltd' in sheet_brand.lower() or 'spector' in sheet_brand.lower() or 'blackstar' in sheet_brand.lower() \
                or 'lag' in sheet_brand.lower() or 'arturia' in sheet_brand.lower() or 'darkglass' in sheet_brand.lower() or 'monterey' in sheet_brand.lower() \
                or 'native instrume' in sheet_brand.lower() or 'native instruments' in sheet_brand.lower() or sheet_brand.lower() == 'ni' or 'rapco' in sheet_brand.lower() \
                or 'rat' in sheet_brand.lower() or 'sequenz' in sheet_brand.lower() or 'source audio' in sheet_brand.lower() or 'sourceaud' in sheet_brand.lower() or 'uxl' in sheet_brand.lower() \
                or 'hamilton' in sheet_brand.lower() or 'enki' in sheet_brand.lower() or 'modal' in sheet_brand.lower() or 'netgear' in sheet_brand.lower() or 'onestage' in sheet_brand.lower() \
                or 'lexicon' in sheet_brand.lower() or 'vox' in sheet_brand.lower():
            if 'CMI' not in sheet_sku:
                sheet_sku = f'{sheet_sku}CMI'
        if 'korg' in sheet_brand.lower():
            if 'CMI' not in sheet_sku:
                sheet_sku = f'{sheet_sku}CMI'
        if sheet_brand.lower() == 'dean':
            if 'DUNIM' not in sheet_sku:
                sheet_sku = f'{sheet_sku}DUNIM'

        if 'crown' in sheet_brand.lower():
            if 'metronome' in sheet_title.lower() or 'piano' in sheet_title.lower() or 'keyboard' in sheet_title.lower() or 'bench' in sheet_title.lower() or 'seat' in sheet_title.lower():

                if 'JD' not in sheet_sku:
                    sheet_sku = f'{sheet_sku}JD'

            else:
                if 'CMI' not in sheet_sku:
                    sheet_sku = f'{sheet_sku}CMI'

        if 'payton' in sheet_brand.lower() or 'schaller' in sheet_brand.lower() or 'jargar' in sheet_brand.lower() or 'kaces' in sheet_brand.lower() or 'tg' in sheet_brand.lower() or 'pickboy' in sheet_brand.lower() \
                or 'fps' in sheet_brand.lower() or 'clarendon' in sheet_brand.lower() or 'johnson' in sheet_brand.lower() or 'teller' in sheet_brand.lower() or 'kun' in sheet_brand.lower():
            if 'PT' not in sheet_sku:
                sheet_sku = f'{sheet_sku}PT'

        if 'pirastro' in sheet_brand.lower():
            if sheet_sku.isdigit():
                if 'PT' not in sheet_sku:
                    sheet_sku = f'{sheet_sku}PT'

        if 'sanchez' in sheet_brand.lower() or 'steinhoff' in sheet_brand.lower() or 'crossfire' in sheet_brand.lower() or 'strauss' in sheet_brand.lower() or 'martinez' in sheet_brand.lower() \
                or 'mooer' in sheet_brand.lower() or 'sonic drive' in sheet_brand.lower() or 'fretz' in sheet_brand.lower() or 'big band' in sheet_brand.lower() or 'casino' in sheet_brand.lower() \
                or 'drumfire' in sheet_brand.lower() or 'handy patch' in sheet_brand.lower() or 'j&d' in sheet_brand.lower() or 'kahzan' in sheet_brand.lower() or 'mojo' \
                in sheet_brand.lower() or 'neowood' in sheet_brand.lower() or 'rare audio' in sheet_brand.lower() or 'slam' in sheet_brand.lower() or 'soundart' in sheet_brand.lower() \
                or 'sound art' in sheet_brand.lower() or 'tiki' in sheet_brand.lower() or 'timberidge' in sheet_brand.lower() or 'tj wilco' in sheet_brand.lower() or 'xhl' in sheet_brand.lower():
            if 'JD' not in sheet_sku:
                sheet_sku = f'{sheet_sku}JD'

        if 'ernie ball' in sheet_brand.lower():
            if 'E' in sheet_sku:
                sheet_sku = sheet_sku.replace('E', '')

            if 'P0' in sheet_sku:
                sheet_sku = sheet_sku.replace('P0', '')

            if 'PO' in sheet_sku:
                sheet_sku = sheet_sku.replace('PO', '')

            if 'CMC' not in sheet_sku:
                sheet_sku = f'{sheet_sku}CMC'

        if 'kyser' in sheet_brand.lower() or 'gruv' in sheet_brand.lower() or 'sterling' in sheet_brand.lower():

            if 'CMC' not in sheet_sku:
                sheet_sku = f'{sheet_sku}CMC'
        if 'hartke' in sheet_brand.lower() or 'loog' in sheet_brand.lower():
            if 'DUNIM' not in sheet_sku:
                sheet_sku = f'{sheet_sku}CMC'
        if 'akai' in sheet_brand.lower() or 'marshall' in sheet_brand.lower() or 'nord' in sheet_brand.lower() or 'hercules' in sheet_brand.lower() \
                or 'auralex' in sheet_brand.lower() or 'alesis' in sheet_brand.lower() or 'samson' in sheet_brand.lower() or 'celestion' in sheet_brand.lower() \
                or 'headrush' in sheet_brand.lower() or 'martin strings' in sheet_brand.lower() or 'm-audio' in sheet_brand.lower() or 'tourtek' in sheet_brand.lower() \
                or 'hardcase' in sheet_brand.lower() or 'mapex' in sheet_brand.lower() or 'udg' in sheet_brand.lower() or 'alto' in sheet_brand.lower():

            if 'EF' not in sheet_sku:
                sheet_sku = f'{sheet_sku}EF'
        if 'elixir' in sheet_brand.lower():
            if '23' in sheet_sku:
                if 'EF' not in sheet_sku:
                    sheet_sku = f'{sheet_sku}EF'
        price = sheet2['D' + str(xx)].value
        price = str(price).replace(',', '')

        #######CHECKING STOCK ON WEBSITES

        stock_avaliability = sheet2['I' + str(xx)].value
        # print(f'stock_avaliability: {stock_avaliability}')

        if stock_avaliability is None:
            continue
        if stock_avaliability.lower() != 'y':
            continue

        if scrapped_items == []:
            scrapped_items.append({'sku': f'{sheet_sku}', 'brand': f'{sheet_brand}', 'price': [price]})
            continue

        #######CHECKING STOCK ON WEBSITES

        for t in scrapped_items:

            if skip == 'true':
                continue

            if t['sku'] == sheet_sku:
                print(t['price'])
                match = 'true'
                t['price'].append(price)

        if match == 'false':
            postage_type = 'false'
            for tt in postage_type_dictionary:
                postage_sku = tt['sku']
                if postage_sku == sheet_sku:
                    postage_type = tt['postage']
                    break

            scrapped_items.append(
                {'sku': f'{sheet_sku}', 'brand': f'{sheet_brand}', 'price': [price], 'postage_type': f'{postage_type}'})
            skip = 'true'

for x in range(1, sheet.max_row + 1):

    try:

        SKU = sheet['A' + str(x)].value

        brand = sheet['D' + str(x)].value
        # try:
        #     if 'gibson' in brand.lower():
        #
        #         pass
        #     else:
        #         continue
        # except:
        #     continue

        if SKU == 'J3102':
            continue

        # if 'TUT21E' in SKU:
        #     pass
        # else:
        #     continue

        ######vvvvvv DEBUG
        # if brand is None:
        #     continue
        # if brand.lower() == 'casio':
        #     pass
        # else:
        #     continue
        #
        # if '450332' in SKU:
        #     print('bing')
        #     print('bong')
        #####^^^^^^^^^^ DEBUG

        if brand is None:
            brand = 'None'
        title = sheet['B' + str(x)].value

        sheet['F' + str(x)].value = SKU

        if SKU in duplicate_values:

            try:

                if brand.lower() == 'daddario' or brand.lower() == 'aquila' or brand.lower() == 'hal leonard' or brand.lower() == 'pro' or brand.lower() == 'koala' or brand.lower() == 'progressive':
                    quantity = 0
                    for g in range(1, sheet.max_row + 1):
                        if sheet['A' + str(g)].value == SKU:
                            quantity += int(sheet['I' + str(g)].value)
                    for g in range(1, sheet.max_row + 1):
                        if sheet['A' + str(g)].value == SKU:
                            sheet['I' + str(g)].value = str(quantity)

            except AttributeError:

                pass

        try:

            if brand.lower() in completed_brands:

                RRP = float(sheet['C' + str(x)].value)
                if 'marshall' in brand.lower():
                        behringer_workbook = openpyxl.load_workbook(rf"Pricing Spreadsheets/Pricing_spreadsheets_supplied_by_suppliers/Marshall.xlsx")
                        behringer_sheet = behringer_workbook['Sheet 1']
                        for items in range(1, behringer_sheet.max_row + 1):
                            behringer_sku = str(behringer_sheet['A' + str(items)].value)
                            behringer_sku = behringer_sku.replace('/','')
                            if behringer_sku is None:
                                continue
                            if behringer_sku in SKU:
                                RRP = str(behringer_sheet['C' + str(items)].value)
                                # RRP = cost.replace('$', '')
                                # RRP = cost.replace(',', '')
                                # RRP = float(cost) * 1.1
                                break
                last_invoiced = str(sheet['O' + str(x)].value)
                is_over_a_year = is_date_over_a_year_ago(last_invoiced)
                quantity = str(sheet['I' + str(x)].value)
                if is_over_a_year is True and float(quantity) >= 1:
                    obsolete_stock = 'Y'
                else:
                    obsolete_stock = 'N'

                if RRP <= 0.01:
                    pass
                else:

                    prices = []
                    prices_a = []
                    prices_b = []
                    postage = ''

                    price, cost = completed_brands[brand.lower()](RRP, title, SKU, obsolete_stock)

                    prices.append(price)
                    prices_a.append(price)
                    prices_b.append(price)

                    min_margin_a = 1.07
                    min_margin_b = 1.15

                    if 'ibanez' in brand.lower():
                        min_margin_a = 1.15
                        min_margin_b = 1.15

                    if RRP < 30:
                        min_sell = float(cost) * 1.15
                        min_margin_a = 1.15
                        min_margin_b = 1.20
                        ###some straps were selling too cheap
                        if brand.lower() == 'ernie ball' and 'strap' in title.lower():
                            min_sell = float(cost) * 1.3
                            min_margin_a = 1.3
                            min_margin_b = 1.3

                    # else:

                    # if 'satchel' in postage.lower() or 'parcel' in postage.lower():
                    #     postage_cost = 7
                    # elif 'label' in postage.lower():
                    #     postage_cost = 3.3
                    # elif 'devil' in postage.lower():
                    #     postage_cost = 1.8
                    # elif 'mini' in postage.lower():
                    #    postage_cost = 0.75
                    # else:
                    #     postage_cost = 1.8
                    # min_sell = (float(cost) * 1.12)+float(postage_cost)

                    # if brand.lower() == 'planet waves':
                    #     min_sell = (float(cost) * 1.20)+ float(postage_cost)

                    if brand.lower() == 'casio':
                        min_sell = 0

                    for all_scrapped_items in scrapped_items:

                        sheet_sku = all_scrapped_items['sku']

                        if sheet_sku == SKU:
                            scrapped_prices = all_scrapped_items['price']
                            try:

                                postage = all_scrapped_items['postage_type']

                                if 'satchel' in postage.lower() or 'parcel' in postage.lower():
                                    postage_cost = 7
                                elif 'label' in postage.lower():
                                    postage_cost = 3.3
                                elif 'devil' in postage.lower():
                                    postage_cost = 2.5
                                elif 'mini' in postage.lower():
                                    postage_cost = 0.75
                                else:
                                    postage_cost = 2.5

                                if 'chinrest' in title.lower() or 'music nomad' in brand.lower():
                                    postage_cost = 7
                                if brand.lower() == 'mahalo' and 'm' in SKU.lower():
                                    postage_cost = 9.5
                                if brand.lower() == 'xtreme' or 'cnb' in brand.lower():
                                    postage_cost = 15
                                if brand.lower() == 'xtreme' and 'tb3' in SKU.lower():
                                    postage_cost = 10
                                if brand.lower() == 'valencia' and 'vc' in SKU.lower():
                                    postage_cost = 20
                                if brand.lower() == 'crossfire' and 'case' in title.lower():
                                    postage_cost = 30
                                if brand.lower() == 'onstage' and 'workstation' in title.lower():
                                    postage_cost = 100
                                if brand.lower() == 'on stage' and 'workstation' in title.lower():
                                    postage_cost = 100
                                if brand.lower() == 'tg' and 'cello case' in title.lower():
                                    postage_cost = 120
                                if 'remo' in brand.lower() and 'head' in title.lower():
                                    postage_cost = 8
                                if brand.lower() == 'v-case':
                                    postage_cost = 30
                                if 'subwoofer' in title.lower():
                                    postage_cost = 120

                                if 'dxp' in brand.lower() and 'drum' in title.lower() and 'kit' in title.lower():
                                    postage_cost = 75
                                if 'dxp' in brand.lower() and 'fusion' in title.lower():
                                    postage_cost = 75
                                if 'dxp' in brand.lower() and 'cocktail' in title.lower() and 'kit' in title.lower():
                                    postage_cost = 75
                                if brand.lower() == 'armour':
                                    if 'case' in title.lower():
                                        postage_cost = 30
                                    if 'bag' in title.lower():
                                        postage_cost = 15

                                min_sell_a = (float(cost) * min_margin_a) + float(postage_cost)
                                min_sell_b = (float(cost) * min_margin_b) + float(postage_cost) + 0.3

                            except:
                                postage = ''
                                postage_cost = 2.5

                            if 'chinrest' in title.lower() or 'music nomad' in brand.lower():
                                postage_cost = 7
                            if brand.lower() == 'mahalo' and 'm' in SKU.lower():
                                postage_cost = 9.5
                            if brand.lower() == 'xtreme' or 'cnb' in brand.lower():
                                postage_cost = 15
                            if brand.lower() == 'xtreme' and 'tb3' in SKU.lower():
                                postage_cost = 10
                            if brand.lower() == 'valencia' and 'vc' in SKU.lower():
                                postage_cost = 20
                            if brand.lower() == 'crossfire' and 'case' in title.lower():
                                postage_cost = 30
                            if brand.lower() == 'onstage' and 'workstation' in title.lower():
                                postage_cost = 100
                            if brand.lower() == 'tg' and 'cello case' in title.lower():
                                postage_cost = 120
                            if brand.lower() == 'v-case':
                                postage_cost = 30
                            if 'remo' in brand.lower() and 'head' in title.lower():
                                postage_cost = 8
                            if 'subwoofer' in title.lower():
                                postage_cost = 120

                            if 'dxp' in brand.lower() and 'drum' in title.lower() and 'kit' in title.lower():
                                postage_cost = 30
                            if 'dxp' in brand.lower() and 'fusion' in title.lower():
                                postage_cost = 30

                            if brand.lower() == 'armour':
                                if 'case' in title.lower():
                                    postage_cost = 30
                                if 'bag' in title.lower():
                                    postage_cost = 15

                            min_sell_a = (float(cost) * min_margin_a) + float(postage_cost)
                            min_sell_b = (float(cost) * min_margin_b) + float(postage_cost) + 0.3

                            for individual_price in scrapped_prices:

                                if individual_price == "N/A":
                                    continue

                                if float(individual_price) > min_sell_a:
                                    prices_a.append(float(individual_price))
                                if float(individual_price) > min_sell_b:
                                    prices_b.append(float(individual_price))

                            if len(prices_a) == 1:
                                if min_sell_a > prices[0]:
                                    prices_a = [min_sell_a]
                            if len(prices_b) == 1:
                                if min_sell_b > prices[0]:
                                    prices_b = [min_sell_b]

                            break
                    ####### For if there were no scrapped items
                    if len(prices_a) == 1:
                        for tt in postage_type_dictionary:
                            postage_sku = tt['sku']
                            if postage_sku == SKU:
                                postage = tt['postage']

                        if 'satchel' in postage.lower() or 'parcel' in postage.lower():
                            postage_cost = 7
                        elif 'label' in postage.lower():
                            postage_cost = 3.3
                        elif 'devil' in postage.lower():
                            postage_cost = 2.5
                        elif 'mini' in postage.lower():
                            postage_cost = 0.75
                        else:
                            postage_cost = 2.5

                        if 'chinrest' in title.lower() or 'music nomad' in brand.lower():
                            postage_cost = 7
                        if brand.lower() == 'mahalo' and 'm' in SKU.lower():
                            postage_cost = 9.5
                        if brand.lower() == 'xtreme' or 'cnb' in brand.lower():
                            postage_cost = 15
                        if brand.lower() == 'xtreme' and 'tb3' in SKU.lower():
                            postage_cost = 10
                        if brand.lower() == 'valencia' and 'vc' in SKU.lower():
                            postage_cost = 20
                        if brand.lower() == 'crossfire' and 'case' in title.lower():
                            postage_cost = 30
                        if brand.lower() == 'v-case':
                            postage_cost = 30
                        if 'remo' in brand.lower() and 'head' in title.lower():
                            postage_cost = 8
                        if 'subwoofer' in title.lower():
                            postage_cost = 120

                        if 'dxp' in brand.lower() and 'drum' in title.lower() and 'kit' in title.lower():
                            postage_cost = 30
                        if 'dxp' in brand.lower() and 'fusion' in title.lower():
                            postage_cost = 30
                        if brand.lower() == 'tg' and 'cello case' in title.lower():
                            postage_cost = 120
                        if brand.lower() == 'armour':
                            if 'case' in title.lower():
                                postage_cost = 30
                            if 'bag' in title.lower():
                                postage_cost = 15

                        min_sell_a = (float(cost) * min_margin_a) + float(postage_cost)

                        if min_sell_a > prices[0]:
                            prices_a = [min_sell_a]

                    if len(prices_b) == 1:
                        for tt in postage_type_dictionary:
                            postage_sku = tt['sku']
                            if postage_sku == SKU:
                                postage = tt['postage']

                        if 'satchel' in postage.lower() or 'parcel' in postage.lower():
                            postage_cost = 7
                        elif 'label' in postage.lower():
                            postage_cost = 3.3
                        elif 'devil' in postage.lower():
                            postage_cost = 2.5
                        elif 'mini' in postage.lower():
                            postage_cost = 0.75
                        else:
                            postage_cost = 2.5

                        if 'chinrest' in title.lower() or 'music nomad' in brand.lower():
                            postage_cost = 7
                        if brand.lower() == 'mahalo' and 'm' in SKU.lower():
                            postage_cost = 9.5
                        if brand.lower() == 'xtreme' or 'cnb' in brand.lower():
                            postage_cost = 15
                        if brand.lower() == 'xtreme' and 'tb3' in SKU.lower():
                            postage_cost = 10
                        if brand.lower() == 'valencia' and 'vc' in SKU.lower():
                            postage_cost = 20
                        if brand.lower() == 'crossfire' and 'case' in title.lower():
                            postage_cost = 30

                        if brand.lower() == 'v-case':
                            postage_cost = 30
                        if 'remo' in brand.lower() and 'head' in title.lower():
                            postage_cost = 8
                        if brand.lower() == 'tg' and 'cello case' in title.lower():
                            postage_cost = 120
                        if 'subwoofer' in title.lower():
                            postage_cost = 120

                        if 'dxp' in brand.lower() and 'drum' in title.lower() and 'kit' in title.lower():
                            postage_cost = 30
                        if 'dxp' in brand.lower() and 'fusion' in title.lower():
                            postage_cost = 30
                        if brand.lower() == 'armour':
                            if 'case' in title.lower():
                                postage_cost = 30
                            if 'bag' in title.lower():
                                postage_cost = 15

                        min_sell_b = (float(cost) * min_margin_b) + float(postage_cost) + 0.3
                        if min_sell_b > prices[0]:
                            prices_b = [min_sell_b]

                    final_price_a = min(prices_a)
                    final_price_b = min(prices_b)

                    # if float(final_price) < 30:
                    #
                    #     if postage is None:
                    #         postage = ''
                    #     #lines 1108 to 1133 might be redundant from the lines added recently (1060-1070)
                    #     if 'satchel' in postage.lower() or 'parcel' in postage.lower():
                    #         if float(final_price) - float(cost) < 7:
                    #             final_price = cost + 8
                    #     if 'devil' in postage.lower():
                    #         if float(final_price) - float(cost) < 4:
                    #             final_price = cost + 4
                    #     if 'mini' in postage.lower():
                    #         if float(final_price) - float(cost) < 1.5:
                    #             final_price = cost + 1.5
                    #
                    #     price_a = math.ceil(float(final_price)) - 0.05
                    #
                    # elif 30 <= float(final_price) < 100:
                    #
                    #     if float(final_price)-float(cost) < 10:
                    #         final_price = cost+10
                    #     price_a = math.ceil(float(final_price))- 1.05
                    #
                    # elif float(final_price) >=100:
                    #     minimum_differential = float(cost)*0.1
                    #     if float(final_price) - float(cost) < float(minimum_differential):
                    #         final_price = cost + minimum_differential
                    #

                    if float(final_price_a) < 50:
                        price_a = math.ceil(float(final_price_a)) - 0.05
                    else:
                        price_a = math.ceil(float(final_price_a)) - 1.05

                    if float(final_price_b) < 50:
                        price_b = math.ceil(float(final_price_b)) - 0.05
                    else:
                        price_b = math.ceil(float(final_price_b)) - 1.05

                    if obsolete_stock == 'Y':
                        if RRP > 50:
                            price_a = cost * 1.06
                            price_b = cost * 1.16  # Additional 10% for eBay fees
                            final_price_a = cost * 1.06
                            final_price_b = cost * 1.16  # Additional 10% for eBay fees
                        else:
                            # Keep existing logic for items with RRP <= 50
                            price_b = price_a * 1.05

                    ##### Price exceptions #####

                    # if brand.lower() == 'planet waves' and RRP < 30 and final_price < RRP:
                    #     final_price = RRP
                    #     price_a = float(RRP)-0.05

                    if RRP < 15 and final_price_a < RRP:
                        final_price_a = RRP+3
                        price_a = RRP+3

                    if RRP < 15 and final_price_b < RRP:
                        final_price_b = RRP+3
                        price_b = RRP+3

                    if RRP < 15 and price_a < 15 and 'mano' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 15 and price_b < 15 and 'mano' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 10 and price_a < 10 and 'crossfire' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 10 and price_b < 10 and 'crossfire' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 15 and price_a < 15 and 'gotoh' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 15 and price_b < 15 and 'gotoh' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'heritage recor' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'heritage recor' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'superslick' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'superslick' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and brand.lower() == 'pro':
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and brand.lower() == 'pro':
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and brand.lower() == 'cherub':
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and brand.lower() == 'cherub':
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and brand.lower() == 'gt':
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and brand.lower() == 'gt':
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'gt part' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'gt part' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'boomwhacker' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'boomwhacker' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'drumfire' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'drumfire' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'axl' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'axl' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 50 and price_a < 50 and 'toca' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 50 and price_b < 50 and 'toca' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 50 and price_a < 50 and 'opus' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 50 and price_b < 50 and 'opus' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'perris' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'perris' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'percussion plus' in brand.lower():
                        price_a = RRP + 10
                        final_price_a = RRP + 10

                    if RRP < 30 and price_b < 30 and 'percussion plus' in brand.lower():
                        price_b = RRP + 10
                        final_price_b = RRP + 10

                    if RRP < 30 and price_a < 30 and 'peace' in brand.lower():
                        price_a = RRP + 12
                        final_price_a = RRP + 12

                    if RRP < 30 and price_b < 30 and 'peace' in brand.lower():
                        price_b = RRP + 12
                        final_price_b = RRP + 12

                    if RRP < 30 and price_a < 30 and 'slam' in brand.lower():
                        price_a = RRP + 12
                        final_price_a = RRP + 12

                    if RRP < 30 and price_b < 30 and 'slam' in brand.lower():
                        price_b = RRP + 12
                        final_price_b = RRP + 12

                    if RRP < 30 and price_a < 30 and 'leem' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'leem' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 20 and price_a < 20 and 'payton' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 20 and price_b < 20 and 'payton' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 20 and price_a < 20 and 'gibraltar' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 10

                    if RRP < 20 and price_b < 20 and 'gibraltar' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 20 and price_a < 20 and 'maestro' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 10

                    if RRP < 20 and price_b < 20 and 'maestro' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 20 and price_a < 20 and 'hosa' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 20 and price_b < 20 and 'hosa' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 20 and price_a < 20 and 'wittner' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 20 and price_b < 20 and 'wittner' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 20 and price_a < 20 and 'j michael' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 10

                    if RRP < 20 and price_b < 20 and 'j michael' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'fretz' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'fretz' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'martinez' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'martinez' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if RRP < 30 and price_a < 30 and 'wolf' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if RRP < 30 and price_b < 30 and 'wolf' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    if 'learn to play' in brand.lower() or 'koala' in brand.lower():
                        price_a = RRP + 5
                        final_price_a = RRP + 5

                    if 'learn to play' in brand.lower() or 'koala' in brand.lower():
                        price_b = RRP + 5
                        final_price_b = RRP + 5

                    # if SKU == 'XSAPB1047':
                    #     price_a = 24.99
                    # if SKU == 'XSAPB1047-12':
                    #     price_a = 37.99
                    # if SKU == 'XSAPB1152':
                    #     price_a = 24.99
                    # if SKU == 'XSAPB1253':
                    #     price_a = 24.99
                    # if SKU == 'XSAPB1356':
                    #     price_a = 24.99
                    # if SKU == 'XSE0942':
                    #     price_a = 19.99
                    # if SKU == 'XSE0946':
                    #     price_a = 19.99
                    # if SKU == 'XSE1046':
                    #     price_a = 19.99
                    # if SKU == '':
                    #     price_a = 19.99
                    # if SKU == 'XSE1052':
                    #     price_a = 19.99
                    # if SKU == 'XSE1149':
                    #     price_a = 19.99

                    if SKU == '16NITROMESHEF':
                        price_a = 649
                        price_b = 669
                        final_price_b = 669

                    if SKU == '8500175AUSTRALIS':
                        price_a = 199
                        price_b = 229
                        final_price_b = 229

                    if SKU == 'AP270BK':
                        price_a = 1429
                        price_b = 1549
                        final_price_b = 1549

                    if SKU == 'KDP120R':
                        price_a = 1399
                        price_b = 1499
                        final_price_b = 1499
                    if SKU == 'KDP120ES':
                        price_a = 1399
                        price_b = 1499
                        final_price_b = 1499

                    if SKU == 'LEC-VOXPACKCMI':
                        price_a = 499
                        price_b = 499
                        final_price_b = 499

                    if SKU == 'MEA-17NCMI':
                        price_a = 99
                        price_b = 99
                        final_price_b = 99

                    if SKU == 'CDPS110BK':
                        price_a = 549
                        price_b = 549
                        final_price_b = 549
                    if SKU == 'CDPS110WE':
                        price_a = 549
                        price_b = 549
                        final_price_b = 549

                    if SKU == 'LAG-T88DCMI':
                        price_a = 299
                        price_b = 299
                        final_price_b = 299

                    if SKU == '57TSSUB15EF':
                        price_a = 897.95
                        price_b = 949
                        final_price_b = 949



                    price_d = math.ceil(float(final_price_b) * 1.1) - 0.05
                    margin = ((float(price_a) - float(cost)) / float(cost)) * 100
                    # print(f'Margin: {margin}')
                    price_c = math.ceil(float(final_price_b) * 1.05) - 0.05
                    price_e = math.ceil(float(final_price_b) * 1.12) - 0.05

                    # Determine category for clearance items
                    if obsolete_stock == 'Y' and RRP > 50:
                        category = 'Clearance Sale'
                    else:
                        category = ''

                    final_worksheet.append(
                        [SKU, brand, title, RRP, price_a, price_b, price_c, price_d, cost, str(margin), postage,
                         price_e, obsolete_stock, category])

                    print(
                        f'\n{SKU} repriced:\nRRP: {RRP}\nPrice A/C: {price_a}\nPrice B: {price_b}\nPrice D: {price_d}\n')




        except Exception as e:

            print('Error found. \n\n')

            print(repr(e))
            pass



    except TypeError:
        sheet.delete_rows(int(x), 2)
        continue

final_workbook.save("Pricing Spreadsheets/Master_Pricing_Speadsheet/Final_pricing_spreadsheet.xlsx")

read_file = pd.read_excel("Pricing Spreadsheets/Master_Pricing_Speadsheet/Final_pricing_spreadsheet.xlsx")
read_file.to_csv("Pricing Spreadsheets/Master_Pricing_Speadsheet/Final_pricing_spreadsheet.csv",
                 header=True, index=False)

# Upload final file to FTP
try:
    session = ftplib.FTP('ftp.drivehq.com', 'kyaldabomb', os.environ.get('FTP_PASSWORD'))
    with open("Pricing Spreadsheets/Master_Pricing_Speadsheet/Final_pricing_spreadsheet.csv", 'rb') as file:
        session.storbinary('STOR Final_pricing_spreadsheet.csv', file)
    file.close()
    session.quit()
    print("File uploaded to FTP successfully")
except Exception as e:
    print(f"Error uploading to FTP: {str(e)}")
