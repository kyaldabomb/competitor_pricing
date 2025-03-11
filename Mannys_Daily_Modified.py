import openpyxl
import requests
from bs4 import BeautifulSoup
import pprint, time, math, os, traceback, sys, argparse
from datetime import datetime, timedelta

from selenium import webdriver
from selenium_stealth import stealth
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Parse command line arguments
parser = argparse.ArgumentParser(description='Run daily scraper for Mannys')
parser.add_argument('scraper', nargs='?', default='mannys_daily', 
                    help='Scraper name from config')
args = parser.parse_args()

# Email notification function
def send_email_notification(success, items_count=0, error_msg=""):
    print("Sending email notification...")
    try:
        # Email settings
        sender = "kyal@scarlettmusic.com.au"
        receiver = "kyal@scarlettmusic.com.au"
        password = os.environ.get('EMAIL_PASSWORD')
        if not password:
            print("Email password not found in environment variables")
            return
            
        host = "mail.scarlettmusic.com.au"
        port = 587
        
        msg = MIMEMultipart()
        msg['From'] = sender
        msg['To'] = receiver
        
        if success:
            msg['Subject'] = f"Mannys Daily Scraper Success: {items_count} items updated"
            body = f"The Mannys daily web scraper ran successfully and updated {items_count} items."
        else:
            msg['Subject'] = "Mannys Daily Scraper Failed"
            body = f"The Mannys daily web scraper encountered an error:\n\n{error_msg}"
        
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(host, port)
        server.starttls()
        server.login(sender, password)
        server.send_message(msg)
        server.quit()
        print("Email notification sent successfully")
    except Exception as e:
        print(f"Failed to send email notification: {str(e)}")
        print(traceback.format_exc())

# Setup Chrome options for headless operation in GitHub Actions
options = webdriver.ChromeOptions()
options.add_argument("--headless")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)

try:
    # Initialize WebDriver using webdriver-manager
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    # Print Chrome and ChromeDriver version for debugging
    print(f"Chrome version: {driver.capabilities['browserVersion']}")
    print(f"ChromeDriver version: {driver.capabilities['chrome']['chromedriverVersion'].split(' ')[0]}")
    
    stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
            )
    
    # Use local path instead of network path
    file_path = "Pricing Spreadsheets/Mannys.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet']
    
    item_number = 0
    items_scrapped = 0
    
    print(f"Starting to process {sheet.max_row-1} items")
    
    for sheet_line in range(2, sheet.max_row+1):
        try:
            item_number += 1
            
            # Check if we need to scrape this item (based on last scrape date)
            time_last_scrapped = sheet['H' + str(sheet_line)].value
            if time_last_scrapped:
                try:
                    string_datetime_conversion = datetime.strptime(time_last_scrapped, '%m %d %Y')
                    if string_datetime_conversion + timedelta(days=7) > datetime.today():
                        print(f'Item {str(item_number)} scrapped less than 7 days ago, skipping...')
                        continue
                except Exception as date_error:
                    print(f"Error parsing date {time_last_scrapped}: {str(date_error)}")
            
            # Get the URL and fetch the page
            url = sheet['E'+str(sheet_line)].value
            if not url:
                print(f"No URL for item {item_number}, skipping")
                continue
                
            print(f"Processing item {item_number}: {url}")
            
            # Add retry logic for resilience
            max_retries = 3
            for retry in range(max_retries):
                try:
                    r = driver.get(url)
                    
                    # Wait for page to load - look for a common element that should be present
                    try:
                        # Wait up to 10 seconds for the page to load
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.TAG_NAME, "body"))
                        )
                        
                        # Wait for price element specifically (could be any of these classes)
                        WebDriverWait(driver, 10).until(
                            lambda d: d.find_element(By.CLASS_NAME, "selling-price") or 
                                    d.find_element(By.CLASS_NAME, "item-price") or 
                                    d.find_element(By.CLASS_NAME, "price")
                        )
                        
                        # To ensure JavaScript has time to execute
                        time.sleep(2)
                    except Exception as wait_error:
                        print(f"Wait error: {str(wait_error)}")
                        # Continue anyway - we'll handle missing elements later
                    
                    break
                except Exception as e:
                    if retry == max_retries - 1:
                        raise
                    print(f"Retry {retry+1}/{max_retries} for {url}: {str(e)}")
                    time.sleep(5)
            
            # Get full page source after everything has loaded
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            
            # Extract price with multiple approaches
            price = "N/A"
            price_found = False
            
            # Approach 1: Try using selling-price class
            price_element = soup.find(class_='selling-price')
            if price_element:
                try:
                    price = price_element.text.strip()
                    price = price.replace('
            
            # Extract description with multiple approaches
            description = 'N/A'
            description_found = False
            
            # Approach 1: Try text-cutoff-wrap class
            try:
                description_div = soup.find('div', class_='text-cutoff-wrap')
                if description_div:
                    description = description_div.text.strip()
                    description_found = True
                    print("Found description using text-cutoff-wrap")
            except Exception as e:
                print(f"Error finding description in text-cutoff-wrap: {str(e)}")
            
            # Approach 2: Try productInfo-content class
            if not description_found:
                try:
                    description_div = soup.find(class_='productInfo-content')
                    if description_div:
                        description = description_div.text.strip()
                        description_found = True
                        print("Found description using productInfo-content")
                except Exception as e:
                    print(f"Error finding description in productInfo-content: {str(e)}")
                
            # Approach 3: Try product-description class
            if not description_found:
                try:
                    description_div = soup.find(class_='product-description')
                    if description_div:
                        description = description_div.text.strip()
                        description_found = True
                        print("Found description using product-description")
                except Exception as e:
                    print(f"Error finding description in product-description: {str(e)}")
            
            # Approach 4: Try text-block class
            if not description_found:
                try:
                    description_div = soup.find(class_='text-block')
                    if description_div:
                        description = description_div.text.strip()
                        description_found = True
                        print("Found description using text-block")
                except Exception as e:
                    print(f"Error finding description in text-block: {str(e)}")
                    
            if not description_found:
                print(f"Could not find description for {url} using any method")
            
            # Extract image
            try:
                image_element = soup.find(class_='product-detail-img')
                if image_element and 'src' in image_element.attrs:
                    image = f"https://www.mannys.com.au/{image_element['src']}"
                else:
                    image = "N/A"
            except Exception as e:
                image = "N/A"
                print(f"Error getting image: {str(e)}")
            
            # Check stock availability with multiple approaches
            stock_avaliable = 'n'
            
            # Approach 1: Try online-stock class
            try:
                stock = soup.find(class_='online-stock')
                if stock and stock.text and ('In Stock' in stock.text or 'Low Stock' in stock.text):
                    stock_avaliable = 'y'
                    print("Found stock using online-stock class")
            except Exception as e:
                print(f"Error checking online-stock: {str(e)}")
            
            # Approach 2: Try online-stock-status class
            if stock_avaliable == 'n':
                try:
                    stock = soup.find(class_='online-stock-status')
                    if stock and stock.text and ('In Stock' in stock.text or 'Low Stock' in stock.text):
                        stock_avaliable = 'y'
                        print("Found stock using online-stock-status class")
                except Exception as e:
                    print(f"Error checking online-stock-status: {str(e)}")
            
            # Approach 3: Try online-stock-green class
            if stock_avaliable == 'n':
                try:
                    stock = soup.find(class_='online-stock-green')
                    if stock and stock.text:
                        stock_avaliable = 'y'
                        print("Found stock using online-stock-green class")
                except Exception as e:
                    print(f"Error checking online-stock-green: {str(e)}")
                
            # Approach 4: Look for any stock indicator in the page
            if stock_avaliable == 'n':
                try:
                    # Find all elements containing stock-related text
                    stock_elements = [tag for tag in soup.find_all() if tag.string and 
                                     ('In Stock' in tag.string or 'Low Stock' in tag.string)]
                    if stock_elements:
                        stock_avaliable = 'y'
                        print("Found stock using text search")
                except Exception as e:
                    print(f"Error finding stock with text search: {str(e)}")
                    
            # Approach 5: Using JavaScript
            if stock_avaliable == 'n':
                try:
                    has_stock = driver.execute_script("""
                        var page = document.body.innerHTML;
                        return page.includes('In Stock') || page.includes('Low Stock');
                    """)
                    if has_stock:
                        stock_avaliable = 'y'
                        print("Found stock using JavaScript")
                except Exception as e:
                    print(f"Error checking stock with JavaScript: {str(e)}")
                    
            print(f"Stock availability: {stock_avaliable}")
            
            # Get current date
            today = datetime.now()
            date = today.strftime('%m %d %Y')
            
            # Update the Excel sheet
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliable
            
            items_scrapped += 1
            print(f'Item {str(item_number)} scraped successfully')
            
            # Add a pause to be gentle with the server
            time.sleep(3)
            
            # Save periodically
            if int(items_scrapped) % 10 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(file_path)
                    print("Sheet saved successfully")
                except Exception as e:
                    print(f"Error occurred while saving the Excel file: {str(e)}")
        
        except Exception as item_error:
            print(f"Error processing item {item_number}: {str(item_error)}")
            print(traceback.format_exc())
            # Continue with next item even if this one fails
    
    # Final save
    wb.save(file_path)
    print(f"Scraping completed successfully. Updated {items_scrapped} items.")
    send_email_notification(True, items_scrapped)
    
except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")
    
    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")
    
    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)  # Exit with error code
finally:
    # Always close the driver
    try:
        driver.quit()
    except:
        pass
, '').replace('\n', '').replace(',', '')
                    price_found = True
                    print(f"Found price using selling-price class: {price}")
                except Exception as e:
                    print(f"Error extracting price from selling-price: {str(e)}")
            
            # Approach 2: Try using item-price class
            if not price_found:
                price_element = soup.find(class_='item-price')
                if price_element:
                    try:
                        price = price_element.text.strip()
                        price = price.replace('
            
            # Extract description
            try:
                # Look for description in the new HTML structure
                description_div = soup.find('div', class_='text-cutoff-wrap')
                if description_div:
                    description = description_div.text.strip()
                else:
                    # Fallback to previous structure
                    description = soup.find(class_='productInfo-content').text
            except:
                description = 'N/A'
                print(f"Could not find description for {url}")
            
            # Extract image
            try:
                image_element = soup.find(class_='product-detail-img')
                if image_element and 'src' in image_element.attrs:
                    image = f"https://www.mannys.com.au/{image_element['src']}"
                else:
                    image = "N/A"
            except Exception as e:
                image = "N/A"
                print(f"Error getting image: {str(e)}")
            
            # Check stock availability
            stock_avaliable = 'n'
            try:
                # First try the new structure from the page you shared
                stock = soup.find(class_='online-stock')
                if stock and ('In Stock' in stock.text or 'Low Stock' in stock.text):
                    stock_avaliable = 'y'
            except:
                # Fallback to the previous structure
                try:
                    stock = soup.find(class_='online-stock-status in-stock')
                    if stock and 'In Stock' in stock.text:
                        stock_avaliable = 'y'
                except:
                    try:
                        stock = soup.find(class_='online-stock-statusin-stock')
                        if stock:
                            stock_avaliable = 'y'
                    except:
                        # One more attempt with a more general approach
                        stock_elements = soup.find_all(class_=lambda c: c and 'stock' in c.lower())
                        for elem in stock_elements:
                            if 'In Stock' in elem.text or 'Low Stock' in elem.text:
                                stock_avaliable = 'y'
                                break
            
            # Get current date
            today = datetime.now()
            date = today.strftime('%m %d %Y')
            
            # Update the Excel sheet
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliable
            
            items_scrapped += 1
            print(f'Item {str(item_number)} scraped successfully')
            
            # Add a pause to be gentle with the server
            time.sleep(3)
            
            # Save periodically
            if int(items_scrapped) % 10 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(file_path)
                    print("Sheet saved successfully")
                except Exception as e:
                    print(f"Error occurred while saving the Excel file: {str(e)}")
        
        except Exception as item_error:
            print(f"Error processing item {item_number}: {str(item_error)}")
            print(traceback.format_exc())
            # Continue with next item even if this one fails
    
    # Final save
    wb.save(file_path)
    print(f"Scraping completed successfully. Updated {items_scrapped} items.")
    send_email_notification(True, items_scrapped)
    
except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")
    
    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")
    
    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)  # Exit with error code
finally:
    # Always close the driver
    try:
        driver.quit()
    except:
        pass
, '').replace('\n', '').replace(',', '')
                        price_found = True
                        print(f"Found price using item-price class: {price}")
                    except Exception as e:
                        print(f"Error extracting price from item-price: {str(e)}")
            
            # Approach 3: Try using price class
            if not price_found:
                price_element = soup.find(class_='price')
                if price_element:
                    try:
                        price = price_element.text.strip()
                        price = price.replace('
            
            # Extract description
            try:
                # Look for description in the new HTML structure
                description_div = soup.find('div', class_='text-cutoff-wrap')
                if description_div:
                    description = description_div.text.strip()
                else:
                    # Fallback to previous structure
                    description = soup.find(class_='productInfo-content').text
            except:
                description = 'N/A'
                print(f"Could not find description for {url}")
            
            # Extract image
            try:
                image_element = soup.find(class_='product-detail-img')
                if image_element and 'src' in image_element.attrs:
                    image = f"https://www.mannys.com.au/{image_element['src']}"
                else:
                    image = "N/A"
            except Exception as e:
                image = "N/A"
                print(f"Error getting image: {str(e)}")
            
            # Check stock availability
            stock_avaliable = 'n'
            try:
                # First try the new structure from the page you shared
                stock = soup.find(class_='online-stock')
                if stock and ('In Stock' in stock.text or 'Low Stock' in stock.text):
                    stock_avaliable = 'y'
            except:
                # Fallback to the previous structure
                try:
                    stock = soup.find(class_='online-stock-status in-stock')
                    if stock and 'In Stock' in stock.text:
                        stock_avaliable = 'y'
                except:
                    try:
                        stock = soup.find(class_='online-stock-statusin-stock')
                        if stock:
                            stock_avaliable = 'y'
                    except:
                        # One more attempt with a more general approach
                        stock_elements = soup.find_all(class_=lambda c: c and 'stock' in c.lower())
                        for elem in stock_elements:
                            if 'In Stock' in elem.text or 'Low Stock' in elem.text:
                                stock_avaliable = 'y'
                                break
            
            # Get current date
            today = datetime.now()
            date = today.strftime('%m %d %Y')
            
            # Update the Excel sheet
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliable
            
            items_scrapped += 1
            print(f'Item {str(item_number)} scraped successfully')
            
            # Add a pause to be gentle with the server
            time.sleep(3)
            
            # Save periodically
            if int(items_scrapped) % 10 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(file_path)
                    print("Sheet saved successfully")
                except Exception as e:
                    print(f"Error occurred while saving the Excel file: {str(e)}")
        
        except Exception as item_error:
            print(f"Error processing item {item_number}: {str(item_error)}")
            print(traceback.format_exc())
            # Continue with next item even if this one fails
    
    # Final save
    wb.save(file_path)
    print(f"Scraping completed successfully. Updated {items_scrapped} items.")
    send_email_notification(True, items_scrapped)
    
except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")
    
    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")
    
    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)  # Exit with error code
finally:
    # Always close the driver
    try:
        driver.quit()
    except:
        pass
, '').replace('\n', '').replace(',', '')
                        price_found = True
                        print(f"Found price using price class: {price}")
                    except Exception as e:
                        print(f"Error extracting price from price class: {str(e)}")
            
            # Approach 4: Try JavaScript execution to get price
            if not price_found:
                try:
                    # Execute JavaScript to look for price elements
                    price_js = driver.execute_script("""
                        var priceElements = document.querySelectorAll('.selling-price, .item-price, .price');
                        for (var i = 0; i < priceElements.length; i++) {
                            if (priceElements[i].textContent.includes('
            
            # Extract description
            try:
                # Look for description in the new HTML structure
                description_div = soup.find('div', class_='text-cutoff-wrap')
                if description_div:
                    description = description_div.text.strip()
                else:
                    # Fallback to previous structure
                    description = soup.find(class_='productInfo-content').text
            except:
                description = 'N/A'
                print(f"Could not find description for {url}")
            
            # Extract image
            try:
                image_element = soup.find(class_='product-detail-img')
                if image_element and 'src' in image_element.attrs:
                    image = f"https://www.mannys.com.au/{image_element['src']}"
                else:
                    image = "N/A"
            except Exception as e:
                image = "N/A"
                print(f"Error getting image: {str(e)}")
            
            # Check stock availability
            stock_avaliable = 'n'
            try:
                # First try the new structure from the page you shared
                stock = soup.find(class_='online-stock')
                if stock and ('In Stock' in stock.text or 'Low Stock' in stock.text):
                    stock_avaliable = 'y'
            except:
                # Fallback to the previous structure
                try:
                    stock = soup.find(class_='online-stock-status in-stock')
                    if stock and 'In Stock' in stock.text:
                        stock_avaliable = 'y'
                except:
                    try:
                        stock = soup.find(class_='online-stock-statusin-stock')
                        if stock:
                            stock_avaliable = 'y'
                    except:
                        # One more attempt with a more general approach
                        stock_elements = soup.find_all(class_=lambda c: c and 'stock' in c.lower())
                        for elem in stock_elements:
                            if 'In Stock' in elem.text or 'Low Stock' in elem.text:
                                stock_avaliable = 'y'
                                break
            
            # Get current date
            today = datetime.now()
            date = today.strftime('%m %d %Y')
            
            # Update the Excel sheet
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliable
            
            items_scrapped += 1
            print(f'Item {str(item_number)} scraped successfully')
            
            # Add a pause to be gentle with the server
            time.sleep(3)
            
            # Save periodically
            if int(items_scrapped) % 10 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(file_path)
                    print("Sheet saved successfully")
                except Exception as e:
                    print(f"Error occurred while saving the Excel file: {str(e)}")
        
        except Exception as item_error:
            print(f"Error processing item {item_number}: {str(item_error)}")
            print(traceback.format_exc())
            # Continue with next item even if this one fails
    
    # Final save
    wb.save(file_path)
    print(f"Scraping completed successfully. Updated {items_scrapped} items.")
    send_email_notification(True, items_scrapped)
    
except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")
    
    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")
    
    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)  # Exit with error code
finally:
    # Always close the driver
    try:
        driver.quit()
    except:
        pass
)) {
                                return priceElements[i].textContent.trim();
                            }
                        }
                        return 'N/A';
                    """)
                    if price_js and price_js != 'N/A':
                        price = price_js.replace('
            
            # Extract description
            try:
                # Look for description in the new HTML structure
                description_div = soup.find('div', class_='text-cutoff-wrap')
                if description_div:
                    description = description_div.text.strip()
                else:
                    # Fallback to previous structure
                    description = soup.find(class_='productInfo-content').text
            except:
                description = 'N/A'
                print(f"Could not find description for {url}")
            
            # Extract image
            try:
                image_element = soup.find(class_='product-detail-img')
                if image_element and 'src' in image_element.attrs:
                    image = f"https://www.mannys.com.au/{image_element['src']}"
                else:
                    image = "N/A"
            except Exception as e:
                image = "N/A"
                print(f"Error getting image: {str(e)}")
            
            # Check stock availability
            stock_avaliable = 'n'
            try:
                # First try the new structure from the page you shared
                stock = soup.find(class_='online-stock')
                if stock and ('In Stock' in stock.text or 'Low Stock' in stock.text):
                    stock_avaliable = 'y'
            except:
                # Fallback to the previous structure
                try:
                    stock = soup.find(class_='online-stock-status in-stock')
                    if stock and 'In Stock' in stock.text:
                        stock_avaliable = 'y'
                except:
                    try:
                        stock = soup.find(class_='online-stock-statusin-stock')
                        if stock:
                            stock_avaliable = 'y'
                    except:
                        # One more attempt with a more general approach
                        stock_elements = soup.find_all(class_=lambda c: c and 'stock' in c.lower())
                        for elem in stock_elements:
                            if 'In Stock' in elem.text or 'Low Stock' in elem.text:
                                stock_avaliable = 'y'
                                break
            
            # Get current date
            today = datetime.now()
            date = today.strftime('%m %d %Y')
            
            # Update the Excel sheet
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliable
            
            items_scrapped += 1
            print(f'Item {str(item_number)} scraped successfully')
            
            # Add a pause to be gentle with the server
            time.sleep(3)
            
            # Save periodically
            if int(items_scrapped) % 10 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(file_path)
                    print("Sheet saved successfully")
                except Exception as e:
                    print(f"Error occurred while saving the Excel file: {str(e)}")
        
        except Exception as item_error:
            print(f"Error processing item {item_number}: {str(item_error)}")
            print(traceback.format_exc())
            # Continue with next item even if this one fails
    
    # Final save
    wb.save(file_path)
    print(f"Scraping completed successfully. Updated {items_scrapped} items.")
    send_email_notification(True, items_scrapped)
    
except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")
    
    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")
    
    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)  # Exit with error code
finally:
    # Always close the driver
    try:
        driver.quit()
    except:
        pass
, '').replace('\n', '').replace(',', '')
                        price_found = True
                        print(f"Found price using JavaScript: {price}")
                except Exception as e:
                    print(f"Error executing JavaScript for price: {str(e)}")
            
            # Approach 5: Manual search for any price-like text
            if not price_found:
                try:
                    # Find all elements with "$" in their text
                    possible_price_elements = [tag for tag in soup.find_all() if tag.string and '
            
            # Extract description
            try:
                # Look for description in the new HTML structure
                description_div = soup.find('div', class_='text-cutoff-wrap')
                if description_div:
                    description = description_div.text.strip()
                else:
                    # Fallback to previous structure
                    description = soup.find(class_='productInfo-content').text
            except:
                description = 'N/A'
                print(f"Could not find description for {url}")
            
            # Extract image
            try:
                image_element = soup.find(class_='product-detail-img')
                if image_element and 'src' in image_element.attrs:
                    image = f"https://www.mannys.com.au/{image_element['src']}"
                else:
                    image = "N/A"
            except Exception as e:
                image = "N/A"
                print(f"Error getting image: {str(e)}")
            
            # Check stock availability
            stock_avaliable = 'n'
            try:
                # First try the new structure from the page you shared
                stock = soup.find(class_='online-stock')
                if stock and ('In Stock' in stock.text or 'Low Stock' in stock.text):
                    stock_avaliable = 'y'
            except:
                # Fallback to the previous structure
                try:
                    stock = soup.find(class_='online-stock-status in-stock')
                    if stock and 'In Stock' in stock.text:
                        stock_avaliable = 'y'
                except:
                    try:
                        stock = soup.find(class_='online-stock-statusin-stock')
                        if stock:
                            stock_avaliable = 'y'
                    except:
                        # One more attempt with a more general approach
                        stock_elements = soup.find_all(class_=lambda c: c and 'stock' in c.lower())
                        for elem in stock_elements:
                            if 'In Stock' in elem.text or 'Low Stock' in elem.text:
                                stock_avaliable = 'y'
                                break
            
            # Get current date
            today = datetime.now()
            date = today.strftime('%m %d %Y')
            
            # Update the Excel sheet
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliable
            
            items_scrapped += 1
            print(f'Item {str(item_number)} scraped successfully')
            
            # Add a pause to be gentle with the server
            time.sleep(3)
            
            # Save periodically
            if int(items_scrapped) % 10 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(file_path)
                    print("Sheet saved successfully")
                except Exception as e:
                    print(f"Error occurred while saving the Excel file: {str(e)}")
        
        except Exception as item_error:
            print(f"Error processing item {item_number}: {str(item_error)}")
            print(traceback.format_exc())
            # Continue with next item even if this one fails
    
    # Final save
    wb.save(file_path)
    print(f"Scraping completed successfully. Updated {items_scrapped} items.")
    send_email_notification(True, items_scrapped)
    
except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")
    
    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")
    
    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)  # Exit with error code
finally:
    # Always close the driver
    try:
        driver.quit()
    except:
        pass
 in tag.string]
                    if possible_price_elements:
                        raw_price = possible_price_elements[0].string.strip()
                        # Extract the price using regex if needed
                        import re
                        price_match = re.search(r'\$[\d,]+(?:\.\d+)?', raw_price)
                        if price_match:
                            price = price_match.group(0).replace('
            
            # Extract description
            try:
                # Look for description in the new HTML structure
                description_div = soup.find('div', class_='text-cutoff-wrap')
                if description_div:
                    description = description_div.text.strip()
                else:
                    # Fallback to previous structure
                    description = soup.find(class_='productInfo-content').text
            except:
                description = 'N/A'
                print(f"Could not find description for {url}")
            
            # Extract image
            try:
                image_element = soup.find(class_='product-detail-img')
                if image_element and 'src' in image_element.attrs:
                    image = f"https://www.mannys.com.au/{image_element['src']}"
                else:
                    image = "N/A"
            except Exception as e:
                image = "N/A"
                print(f"Error getting image: {str(e)}")
            
            # Check stock availability
            stock_avaliable = 'n'
            try:
                # First try the new structure from the page you shared
                stock = soup.find(class_='online-stock')
                if stock and ('In Stock' in stock.text or 'Low Stock' in stock.text):
                    stock_avaliable = 'y'
            except:
                # Fallback to the previous structure
                try:
                    stock = soup.find(class_='online-stock-status in-stock')
                    if stock and 'In Stock' in stock.text:
                        stock_avaliable = 'y'
                except:
                    try:
                        stock = soup.find(class_='online-stock-statusin-stock')
                        if stock:
                            stock_avaliable = 'y'
                    except:
                        # One more attempt with a more general approach
                        stock_elements = soup.find_all(class_=lambda c: c and 'stock' in c.lower())
                        for elem in stock_elements:
                            if 'In Stock' in elem.text or 'Low Stock' in elem.text:
                                stock_avaliable = 'y'
                                break
            
            # Get current date
            today = datetime.now()
            date = today.strftime('%m %d %Y')
            
            # Update the Excel sheet
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliable
            
            items_scrapped += 1
            print(f'Item {str(item_number)} scraped successfully')
            
            # Add a pause to be gentle with the server
            time.sleep(3)
            
            # Save periodically
            if int(items_scrapped) % 10 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(file_path)
                    print("Sheet saved successfully")
                except Exception as e:
                    print(f"Error occurred while saving the Excel file: {str(e)}")
        
        except Exception as item_error:
            print(f"Error processing item {item_number}: {str(item_error)}")
            print(traceback.format_exc())
            # Continue with next item even if this one fails
    
    # Final save
    wb.save(file_path)
    print(f"Scraping completed successfully. Updated {items_scrapped} items.")
    send_email_notification(True, items_scrapped)
    
except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")
    
    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")
    
    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)  # Exit with error code
finally:
    # Always close the driver
    try:
        driver.quit()
    except:
        pass
, '').replace(',', '')
                            price_found = True
                            print(f"Found price using text search: {price}")
                except Exception as e:
                    print(f"Error finding price manually: {str(e)}")
            
            if not price_found:
                print(f"Could not find price for {url} using any method")
            
            # Extract description
            try:
                # Look for description in the new HTML structure
                description_div = soup.find('div', class_='text-cutoff-wrap')
                if description_div:
                    description = description_div.text.strip()
                else:
                    # Fallback to previous structure
                    description = soup.find(class_='productInfo-content').text
            except:
                description = 'N/A'
                print(f"Could not find description for {url}")
            
            # Extract image
            try:
                image_element = soup.find(class_='product-detail-img')
                if image_element and 'src' in image_element.attrs:
                    image = f"https://www.mannys.com.au/{image_element['src']}"
                else:
                    image = "N/A"
            except Exception as e:
                image = "N/A"
                print(f"Error getting image: {str(e)}")
            
            # Check stock availability
            stock_avaliable = 'n'
            try:
                # First try the new structure from the page you shared
                stock = soup.find(class_='online-stock')
                if stock and ('In Stock' in stock.text or 'Low Stock' in stock.text):
                    stock_avaliable = 'y'
            except:
                # Fallback to the previous structure
                try:
                    stock = soup.find(class_='online-stock-status in-stock')
                    if stock and 'In Stock' in stock.text:
                        stock_avaliable = 'y'
                except:
                    try:
                        stock = soup.find(class_='online-stock-statusin-stock')
                        if stock:
                            stock_avaliable = 'y'
                    except:
                        # One more attempt with a more general approach
                        stock_elements = soup.find_all(class_=lambda c: c and 'stock' in c.lower())
                        for elem in stock_elements:
                            if 'In Stock' in elem.text or 'Low Stock' in elem.text:
                                stock_avaliable = 'y'
                                break
            
            # Get current date
            today = datetime.now()
            date = today.strftime('%m %d %Y')
            
            # Update the Excel sheet
            sheet['D' + str(sheet_line)].value = price
            sheet['F' + str(sheet_line)].value = image
            sheet['G' + str(sheet_line)].value = description
            sheet['H' + str(sheet_line)].value = date
            sheet['I' + str(sheet_line)].value = stock_avaliable
            
            items_scrapped += 1
            print(f'Item {str(item_number)} scraped successfully')
            
            # Add a pause to be gentle with the server
            time.sleep(3)
            
            # Save periodically
            if int(items_scrapped) % 10 == 0:
                print(f'Saving Sheet... Please wait....')
                try:
                    wb.save(file_path)
                    print("Sheet saved successfully")
                except Exception as e:
                    print(f"Error occurred while saving the Excel file: {str(e)}")
        
        except Exception as item_error:
            print(f"Error processing item {item_number}: {str(item_error)}")
            print(traceback.format_exc())
            # Continue with next item even if this one fails
    
    # Final save
    wb.save(file_path)
    print(f"Scraping completed successfully. Updated {items_scrapped} items.")
    send_email_notification(True, items_scrapped)
    
except Exception as e:
    error_message = str(e)
    full_traceback = traceback.format_exc()
    print(f"Error in scraping: {error_message}")
    print(f"Traceback:\n{full_traceback}")
    
    try:
        wb.save(file_path)
        print("Saved progress before error")
    except Exception as save_error:
        print(f"Could not save progress after error: {str(save_error)}")
    
    send_email_notification(False, error_msg=f"{error_message}\n\nFull traceback:\n{full_traceback}")
    sys.exit(1)  # Exit with error code
finally:
    # Always close the driver
    try:
        driver.quit()
    except:
        pass
